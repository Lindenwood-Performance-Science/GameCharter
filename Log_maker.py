# -*- coding: utf-8 -*-
"""
Created on Sun Dec 10 17:03:35 2023

@author: Bennett Stice
"""
import psycopg2
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import io
import os

##globals
header_pos=['A2','B2','C2','D2','E2','F2','G2','H2','I2','J2','K2','L2','M2','N2','O2','P2','Q2','R2','S2','T2','U2','V2','W2','X2','Y2','Z2','AA2','AB2','AC2','AD2','AE2','AF2']
pitch_headersa=['Date','Opponent','Pitches','Pitches Per Inning','Peak Velocity','1st Pitch % (60)','OS Strike % (50)','S/M % (25)','FB Velo-Range','Chases','A3P % (60)','OPP SLG % (.400)',' WHIP ','OPP OBP','OPP OPS']
pitcher_headersb=['LO % (65)','Overall Strike % (60)','OPP BAA','BAA w/ 2K (.150)','BAA BIP','Freebases','Strikeouts','Pitches Ahead %','Pitches Behind %','Strikeout %','Ground Ball Out %','Fly Ball Out %','PA Win %','Pitch Spread %','Pitch Spread Strike %', 'Pitch Spread Whiff %','Pitch Spread Hit %']
pitcher_headers=pitch_headersa+pitcher_headersb
season_game_headersa=['Name','Pitches','Pitches Per Inning','Peak Velocity','1st Pitch % (60)','OS Strike % (50)','S/M % (25)','FB Velo-Range','Chases','A3P % (60)','OPP SLG % (.400)',' WHIP ','OPP OBP','OPP OPS']
season_game_headersb=['LO % (65)','Overall Strike % (60)','OPP BAA','BAA w/ 2K (.150)','BAA BIP','Freebases','Strikeouts','Pitches Ahead %','Pitches Behind %','Strikeout %','Ground Ball Out %','Fly Ball Out %','PA Win %','Pitch Spread %','Pitch Spread Strike %', 'Pitch Spread Whiff %','Pitch Spread Hit %' ]
season_game_headers=season_game_headersa+season_game_headersb


def insert_header(pos,name,sheetName):
    sheetName[pos]=name
    bold_font = Font(bold=True)
    sheetName[pos].font = bold_font
    
def insert_secondary_header(row_i, col_i, name, sheetName):
    sheetName.cell(row=row_i, column=col_i, value=name)
    bold_font = Font(bold=True)
    sheetName.cell(row=row_i, column=col_i).font = bold_font
    
def wipe(workbooka):
    # Delete all sheets except the active one
    all_sheets = workbooka.sheetnames

    for sheet_name in all_sheets:
        if sheet_name != 'Sheet':
            del workbooka[sheet_name]
            
def create_workbook(file_nameb):
    # Specify the path to the file
    file_path = os.path.join(os.path.expanduser("~"), "OneDrive", "Documents", "Lindenwood Performance Science", "gameCharter", file_nameb)
#   file_path = os.path.join(os.path.expanduser("~"),"LUPS",file_nameb)    

    # Check if the file exists
    if os.path.exists(file_path):
        # Load the existing workbook
        workbook = openpyxl.load_workbook(file_path)
    else:
        # Create a new workbook
        workbook = openpyxl.Workbook()
    
    return_list=[workbook,file_path]
    
    return return_list

def setup(sheetnameb,workbookb,A1title,B1entry,D1title,E1entry,G1title,H1entry,pos,head):
    # Create a new sheet and set it as the active sheet
    new_sheetc = workbookb.create_sheet(sheetnameb)
    workbookb.active = new_sheetc

    new_sheetc['A1']=A1title
    new_sheetc['B1']=B1entry
    bold_font = Font(bold=True)
    new_sheetc['A1'].font = bold_font
    
    new_sheetc['D1']=D1title
    new_sheetc['E1']=E1entry
    new_sheetc['D1'].font = bold_font
    
    new_sheetc['G1']=G1title
    new_sheetc['H1']=H1entry
    new_sheetc['G1'].font = bold_font
    
    for l in range(0,len(head)):
        insert_header(pos[l],head[l],new_sheetc)
        
    new_sheetc['AC1']="FF-CB-SL-CH-SP-FT-CT"
        
    return new_sheetc

def bold_first_column_if_threshold(sheet, threshold):
    for row in range(3, sheet.max_row + 1):
        bold_count = sum(1 for col in sheet.iter_cols(min_row=row, max_row=row) for cell in col if cell.font and cell.font.bold)
        
        if bold_count >= threshold:
            sheet.cell(row=row, column=1).font = Font(bold=True)

def adjust_formating(new_sheetd,row_i):
     
    # Shade cells in the row that have values
    for col_num in range(1, new_sheetd.max_column + 1):
        cell_value = new_sheetd.cell(row=row_i, column=col_num).value
        if cell_value is not None:
            new_sheetd.cell(row=row_i, column=col_num).fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        
    # Adjust column width to fit the widest text entry
    for column in new_sheetd.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            adjusted_width = (max_length + 1)
            new_sheetd.column_dimensions[column[0].column_letter].width = adjusted_width

    # Center all text in the cells
    for row in new_sheetd.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
def savebook(workbookc,file_pathb,endMessage):
    # Close the workbook before saving
    workbookc.close()
   
    # Save the workbook
    workbookc.save(file_pathb)
    
    print(endMessage)
    
def insert_names (cursora,new_sheetb,ending,row_i,col_i, exe):
    ##### Name
    query="SELECT DISTINCT fname, lname FROM pitch_log_t "
    query+=ending
    cursora.execute(query, exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0 ):
        trip=True
    
    for i, (fname, lname) in enumerate(data,3):
        if (trip):
            row_i=i
        full_name = f"{fname} {lname}"
        new_sheetb.cell(row=row_i,column=col_i, value=full_name)

def insert_oppo(cursorb,new_sheetb,ending,row_i,column_i,exe):
    ##### Opponent
    query= "SELECT opponent AS oppo FROM pitch_log_t "
    query+= ending
    cursorb.execute(query,exe)
    data=cursorb.fetchall()
    
    for k, (oppo,) in enumerate(data, 3):
        if (row_i==0):
            row_i=k
        put_in = oppo[0] if isinstance(oppo, tuple) and oppo else oppo
        new_sheetb.cell(row=row_i, column=column_i, value=put_in)
        
def insert_pitches_thrown(cursorb,new_sheetb,ending,row_i,column_i,exe):
    ##### Pitches Thrown
    query="SELECT COUNT(pitch_id) AS pitchCount FROM pitch_log_t "
    query+= ending
    cursorb.execute(query, exe)
    data=cursorb.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    tpc=0 
    for k, (pitchCount,) in enumerate(data, 3):
        if (trip):
            row_i=k
        put_in = int(pitchCount) if pitchCount is not None else 0
        new_sheetb.cell(row=row_i, column=column_i, value=put_in)
        tpc += put_in
    return tpc
        
def insert_pitches_per_inning(cursora,new_sheetb,ending,row_i,col_i,exe,trigger):
    ##### Pitches Per Inning
    if (trigger):       
        query = "SELECT SUM(pitches) AS pitchCount, SUM(max_outs) AS outs FROM (SELECT MAX(pitch_count) as pitches,  "
        query += "MAX(outs_accrued) AS max_outs,fname,lname FROM pitch_log_T WHERE pitch_id <> '0' and opponent <> 'Scrimmage' "
        query += "GROUP BY date, fname, lname) AS max_outs_per_date GROUP BY fname, lname ORDER BY fname, lname"
    else:
        query= "SELECT COUNT(pitch_id) AS pitchCount, MAX(outs_accrued) AS outs FROM pitch_log_t "
        query+=ending
   
    cursora.execute(query, exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    ti=0
    for k, (pitchCount, outs) in enumerate(data, 3):
        if (trip):
            row_i=k
        innings= float(outs)/3 if outs is not None else 0
        ti+=innings 
        count_a=int(pitchCount) if pitchCount is not None else 0
        if (innings!=0):
            put_in = count_a/innings
            put_in=round (put_in,2)
        else:
            put_in=0
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)
    return ti
        
def insert_peak_velo(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Peak Velocity
    query="SELECT MAX(velocity) AS velo FROM pitch_log_t "
    query+=ending
    cursora.execute(query, exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    tpv=0    
    arms=len(data)
    #print(arms)
    for k, (velo,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = int(velo) if velo is not None else 0
        tpv+=put_in 
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)
    return tpv, arms
  
def insert_1st_pitch_strike_percentage(cursora,new_sheetb,ending,row_i,col_i,exe,goodNum):
    ##### 1st Pitch Strike Percentage
    query="SELECT CASE WHEN COUNT(CASE WHEN balls = 0 AND strikes = 0 THEN 1 END) > 0 "
    query+="THEN (COUNT(CASE WHEN balls = 0 AND strikes = 0 AND pitch_result <> 'B' AND pitch_result <> 'HBP' THEN 1 END) * 100.0 / "
    query+="COUNT(CASE WHEN balls = 0 AND strikes = 0 THEN 1 END)) ELSE 0 END AS Percentage FROM pitch_log_t "
    query+= ending
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (percentage,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = int(percentage) if percentage is not None else 0
        cella = new_sheetb.cell(row=row_i, column=col_i, value=put_in)
        
        if put_in >= goodNum:
            cella.font = Font(bold=True)
            
def insert_off_speed_strike_percentage(cursora,new_sheetb,ending,row_i,col_i,exe,goodNum):
    ##### Off-Speed Strike Percentage
    query="SELECT CASE WHEN COUNT(CASE WHEN pitch_type NOT IN ('FF','FT','CT') THEN 1 END) > 0 "
    query+="THEN (COUNT(CASE WHEN pitch_type NOT IN ('FF','FT','CT') AND pitch_result <> 'B' AND pitch_result <> 'HBP' THEN 1 END) * 100.0 / "
    query+="COUNT(CASE WHEN pitch_type NOT IN ('FF','FT','CT') THEN 1 END)) ELSE 0 END AS PercentageOFF FROM pitch_log_t "
    query+=ending
    cursora.execute(query ,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (percentageOFF,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = int(percentageOFF) if percentageOFF is not None else 0
        cella = new_sheetb.cell(row=row_i, column=col_i, value=put_in)
        
        if put_in >= goodNum:
            cella.font = Font(bold=True)
            
def insert_swing_and_miss_percentage(cursora,new_sheetb,ending,row_i,col_i,exe,goodNum):
    ##### Swing and Miss Percentage
    query ="SELECT CASE WHEN COUNT(CASE WHEN pitch_result <>'0' THEN 1 END) > 0 "
    query+="THEN (COUNT(CASE WHEN pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS' THEN 1 END) * 100.0 / "
    query+="COUNT(CASE WHEN pitch_result IN('SS','SSC','D3SS','F','BIP') THEN 1 END)) ELSE 0 END AS Misses FROM pitch_log_t "
    query+= ending
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (misses,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = int(misses) if misses is not None else 0
        cella = new_sheetb.cell(row=row_i, column=col_i, value=put_in)
         
        if put_in >= goodNum:
            cella.font = Font(bold=True)
    
def insert_velo_range(cursora,new_sheetb,ending,row_i,col_i,exe,trigger1,trigger2, trigger3):
    ##### Velocity Range
    query="SELECT MIN(velocity) AS MINV, MAX(velocity) AS MAXV FROM pitch_log_t "
    if trigger1:
        query += "WHERE pitch_type IN ('FF','FT','CT') AND pitch_id <> '0' AND opponent <> 'Scrimmage' GROUP BY fname,lname ORDER BY fname,lname"
    if trigger2:
        query+="WHERE pitch_type IN ('FF','FT','CT') AND date = %s AND opponent = %s AND opponent <> 'Scrimmage' GROUP BY fname,lname ORDER BY fname,lname"
    if trigger3:
        query+="WHERE pitch_type IN ('FF','FT','CT') AND date = %s AND opponent = %s GROUP BY fname,lname ORDER BY fname,lname"
    if not (trigger1 or trigger2 or trigger3):
        query+= ending
        query += " AND pitch_type IN ('FF','FT','CT') "
    
    cursora.execute(query, exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
     
    for k, (minv, maxv) in enumerate(data, 3):
        if (trip):
            row_i=k
        min_value = int(minv) if minv is not None else 0
        max_value = int(maxv) if maxv is not None else 0
        put_in = f"{min_value}-{max_value}"
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)

def insert_chases(cursora,new_sheetb,ending,row_i,col_i,exe,trigger,innings_sub):
    ##### Chases
    if (trigger):
        query = "SELECT SUM(chase_case) AS chases, SUM(max_outs)/3 AS IP FROM (SELECT SUM(CASE WHEN "
        query += "pitch_result='SSC' or pitch_result='D3SS' THEN 1 ELSE 0 END) AS chase_case, MAX(outs_accrued) AS max_outs,fname,lname "
        query += "FROM pitch_log_T WHERE pitch_id <> '0' AND opponent <>'Scrimmage' "
        query += "GROUP BY date, fname, lname) AS max_outs_per_date GROUP BY fname, lname ORDER BY fname, lname"
    else:
        query="SELECT SUM(CASE WHEN pitch_result='SSC' or pitch_result='D3SS' THEN 1 ELSE 0 END) AS chases, Max(outs_accrued)/3 AS IP FROM pitch_log_t "
        query+=ending
    cursora.execute(query, exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True

    for k, (chases,ip,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = int(chases) if chases is not None else 0
        cella=new_sheetb.cell(row=row_i, column=col_i, value=put_in)
        innings= float(ip) if ip is not None else 0
        if (innings_sub!=0):
            innings=innings_sub
        
        if put_in >= innings and put_in !=0:
            cella.font = Font(bold=True)

def insert_ahead_after_3_pitches_percentage(cursora,new_sheetb,ending,row_i,col_i,exe,goodNum):
    ##### Ahead After 3 Pitches Percentage
    query = "SELECT CASE WHEN COUNT(CASE WHEN (balls=1 AND strikes=2) or (balls=2 AND strikes=1) or (balls=0 AND strikes=2 AND pitch_result Not IN ('B','F')) or (balls=2 AND strikes=0 AND pitch_result IN ('B','HBP','BIP'))THEN 1 END) > 0 "
    query+="THEN (COUNT(CASE WHEN (balls=1 AND strikes=2) or (balls=0 AND strikes=2 AND pitch_result Not IN ('B','F')) THEN 1 END) * 100.0 / "
    query+="COUNT(CASE WHEN (balls=1 AND strikes=2) or (balls=2 AND strikes=1) or (balls=0 AND strikes=2 AND pitch_result Not IN ('B','F','HBP')) or (balls=2 AND strikes=0 AND pitch_result IN ('B','HBP','BIP'))THEN 1 END)) ELSE 0 END AS AA3P FROM pitch_log_t "
    query+=ending
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True

    for k, (aa3p,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = int(aa3p) if aa3p is not None else 0
        cella = new_sheetb.cell(row=row_i, column=col_i, value=put_in)
 
        if put_in >= goodNum:
            cella.font = Font(bold=True)   
            
def insert_opponent_slugging_percentage(cursora,new_sheetb,ending,row_i,col_i,exe,goodNum):
    query = "SELECT  COUNT(CASE WHEN bip_result='1B' THEN 1 END) AS Singles , "
    query += "COUNT(CASE WHEN bip_result='2B' THEN 1 END) AS Doubles, "
    query += "COUNT(CASE WHEN bip_result='3B' THEN 1 END) AS Triples, "
    query += "COUNT(CASE WHEN bip_result='HR' THEN 1 END) AS Homeruns, "
    query += "COUNT(CASE WHEN ab_result<>'0' AND pitch_result NOT IN ('B','HBP') THEN 1 END) AS At_Bats "
    query += "FROM pitch_log_T "
    query += ending
    
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (Singles,Doubles,Triples,Homeruns,At_Bats) in enumerate(data,3):
        if (trip):
            row_i=k
        Singles_Val=int(Singles)*1.0 if Singles is not None else 0
        Doubles_Val=int(Doubles)*2.0 if Doubles is not None else 0
        Triples_Val=int(Triples)*3.0 if Triples is not None else 0
        Homeruns_Val=int(Homeruns)*4.0 if Homeruns is not None else 0
        At_Bats=int(At_Bats) if At_Bats is not None else 0
        if At_Bats!=0:
            put_in=round((Singles_Val+Doubles_Val+Triples_Val+Homeruns_Val)/At_Bats,3)
        else:
            put_in=0
        cella=new_sheetb.cell(row=row_i, column=col_i, value=put_in)  
        
        if put_in <= goodNum:
            cella.font = Font(bold=True)
            
def insert_WHIP(cursora,new_sheetb,ending,row_i,col_i,exe,trigger1,trigger2,innings_sub):
    #### Walks and Hits Per Innings Pitched
    if (trigger1):
        query = "SELECT SUM(safe_case) AS safes, SUM(max_outs) AS OUTS FROM (SELECT SUM(CASE WHEN "
        query += "ab_result='safe' AND bip_result <>'E' AND pitch_result <>'HBP' THEN 1 ELSE 0 END) AS safe_case, MAX(outs_accrued) AS max_outs,fname,lname "
        query += "FROM pitch_log_T WHERE pitch_id <> '0' AND opponent <>'Scrimmage' "
        query += "GROUP BY date, fname, lname) AS max_outs_per_date GROUP BY fname, lname ORDER BY fname, lname"
    else:
        query = "SELECT "
        query += "COUNT(CASE WHEN ab_result='safe' AND bip_result <>'E' AND pitch_result <>'HBP' THEN 1 END) AS Safes, "
        query += "MAX(outs_accrued) AS OUTS FROM pitch_log_T  "
        query += ending
        
    if (trigger2):
        query = "SELECT "
        query += "COUNT(CASE WHEN ab_result='safe' AND bip_result <>'E' AND pitch_result <>'HBP' THEN 1 END) AS Safes, "
        query += "MAX(outs_accrued) AS OUTS FROM pitch_log_T WHERE pitch_id<>'0' and opponent <> 'Scrimmage' "
        
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
        
    for k, (Safes,Outs) in enumerate(data,3):
        if (trip):
            row_i=k
        Outs=Outs if Outs is not None else 0
        Safes=int(Safes)*1.0 if Safes is not None else 0
        IP=Outs/3
        
        if (innings_sub!=0):
            IP=innings_sub
        if Outs!=0:
            put_in=round(Safes/IP,3)
        else:
            put_in="∞"
        
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)  

def insert_OBP(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Oppenent On Base Percentage
    query = "SELECT COUNT(CASE WHEN ab_result<>'0' THEN 1 END) AS PA, "
    query+="COUNT(CASE WHEN ab_result='safe'  THEN 1 END) AS SAFES FROM pitch_log_T "
    query+= ending
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
        
    for k, (PA,Safes) in enumerate(data,3):
        if (trip):
            row_i=k
        PA=int(PA) if PA is not None else 0
        Safes=int(Safes)*1.0 if Safes is not None else 0
        if PA!=0:
            put_in=Safes/PA
        else:
            put_in=0
        put_in=round(put_in,3)
        new_sheetb.cell(row=row_i, column=col_i, value=put_in) 
        
def insert_OPS(cursora,new_sheetb,ending,row_i,col_i,exe):
    query = "SELECT  COUNT(CASE WHEN bip_result='1B' THEN 1 END) AS Singles , "
    query += "COUNT(CASE WHEN bip_result='2B' THEN 1 END) AS Doubles, "
    query += "COUNT(CASE WHEN bip_result='3B' THEN 1 END) AS Triples, "
    query += "COUNT(CASE WHEN bip_result='HR' THEN 1 END) AS Homeruns, "
    query += "COUNT(CASE WHEN ab_result<>'0' AND pitch_result NOT IN ('B','HBP') THEN 1 END) AS At_Bats, "
    query += "COUNT(CASE WHEN ab_result<>'0' THEN 1 END) AS PA, "
    query += "COUNT(CASE WHEN ab_result='safe'  THEN 1 END) AS SAFES "
    query += "FROM pitch_log_T "
    query += ending
    
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (Singles,Doubles,Triples,Homeruns,At_Bats,PA,Safes) in enumerate(data,3):
        if (trip):
            row_i=k
        Singles_Val=int(Singles)*1.0 if Singles is not None else 0
        Doubles_Val=int(Doubles)*2.0 if Doubles is not None else 0
        Triples_Val=int(Triples)*3.0 if Triples is not None else 0
        Homeruns_Val=int(Homeruns)*4.0 if Homeruns is not None else 0
        At_Bats=int(At_Bats) if At_Bats is not None else 0
        PA=float(PA) if PA is not None else 0
        Safes=int(Safes)*1.0 if Safes is not None else 0
        if PA!=0:
            OBP=round(Safes/PA,3)
        else:
            OBP=0
        if At_Bats!=0:
            SLG=round((Singles_Val+Doubles_Val+Triples_Val+Homeruns_Val)/At_Bats,3)
        else:
            SLG=0
        
        put_in=OBP+SLG
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)  
    
    
            
def insert_lead_off_out_percentage(cursora,new_sheetb,ending,row_i,col_i,exe,goodNum):
    ##### Lead Off Out Percentage
    query = "SELECT CASE WHEN COUNT(CASE WHEN batter_of_inning = 1 and ab_result <> '0' THEN 1 END) > 0 "
    query+= "THEN (COUNT(CASE WHEN batter_of_inning = 1 AND ab_result = 'out' THEN 1 END) * 100.0 / COUNT(CASE WHEN batter_of_inning = 1 AND (ab_result = 'out' or ab_result = 'safe') THEN 1 END)) ELSE 0 END AS LOO FROM pitch_log_t "
    query+= ending
    cursora.execute(query, exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (loo,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = int(loo) if loo is not None else 0
        cella=new_sheetb.cell(row=row_i, column=col_i, value=put_in)
        
        if put_in >= goodNum:
            cella.font = Font(bold=True)

def insert_overall_strike_percentage(cursora,new_sheetb,ending,row_i,col_i,exe,goodNum):
    ##### Overall Strike Percentage
    query="SELECT CASE WHEN COUNT(CASE WHEN pitch_result <> '0' THEN 1 END) > 0 "
    query+="THEN (COUNT(CASE WHEN pitch_result <> 'B' AND pitch_result <> 'HBP' THEN 1 END) * 100.0 / "
    query+="COUNT(CASE WHEN pitch_result <> '0' THEN 1 END)) ELSE 0 END AS OvePer FROM pitch_log_t "
    query+=ending
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (oveper,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = int(oveper) if oveper is not None else 0
        cella = new_sheetb.cell(row=row_i, column=col_i, value=put_in)
        
        if put_in >= goodNum:
            cella.font = Font(bold=True)

def insert_baa(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Oppenent Batting Average
    query = "SELECT COUNT(CASE WHEN pitch_result<>'B' AND pitch_result<>'HBP' AND bip_result <>'E' AND ab_result<>'0' THEN 1 END) AS AB, "
    query+="COUNT(CASE WHEN bip_result IN ('1B','2B','3B','HR')  THEN 1 END) AS SAFE FROM pitch_log_t "
    query+=ending
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
        
    for k, (AB,Safe) in enumerate(data,3):
        if (trip):
            row_i=k
        AB=int(AB) if AB is not None else 0
        Safe=int(Safe)*1.0 if Safe is not None else 0
        if AB!=0:
            put_in=Safe/AB
        else:
            put_in=0
        put_in=round(put_in,3)
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)  


def insert_baa_with_2_strikes(cursora,new_sheetb,ending,row_i,col_i,exe,goodNum):
    ##### Opponent Batting Average with 2 Strikes
    query="SELECT CASE WHEN COUNT(CASE WHEN strikes = 2 AND ab_result <> '0' and (pitch_result <> 'B' and pitch_result <> 'HBP') THEN 1 END) > 0 "
    query+="THEN (COUNT(CASE WHEN strikes = 2 AND ab_result = 'safe' and bip_result In ('1B','2B','3B','HR') and bip_result <> 'E' THEN 1 END) * 1.0 / "
    query+="COUNT(CASE WHEN strikes = 2 AND ab_result <> '0' and (pitch_result <> 'B' and pitch_result <> 'HBP') THEN 1 END)) ELSE 0 END AS BAAw2K FROM pitch_log_t "
    query+= ending
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (baaw2k,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = round(float(baaw2k),3) if baaw2k is not None else 0
        cella = new_sheetb.cell(row=row_i, column=col_i, value=put_in)
        
        if put_in <=goodNum:
            cella.font = Font(bold=True)
            
def insert_baa_bip(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Oppenent Batting Average on Balls in Play
    query = "SELECT COUNT(CASE WHEN pitch_result = 'BIP' THEN 1 END) AS BIP, "
    query+="COUNT(CASE WHEN ab_result = 'safe' AND pitch_result = 'BIP' THEN 1 END) AS BIPSAFE FROM pitch_log_t "
    query += ending
    cursora.execute(query, exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (bip,bipsafe) in enumerate(data,3):
        if (trip):
            row_i=k
        bip=int(bip) if bip is not None else 0
        bipsafe=int(bipsafe)*1.0 if bipsafe is not None else 0
        if bip!=0:
            put_in=bipsafe/bip
        else:
            put_in=0
        put_in=round(put_in,3)
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)  
        
            
def insert_freebases_count(cursora,new_sheetb,ending,row_i,col_i,exe):
    ###### Count of Walks and Hit by Pitches
    query = "SELECT COUNT(CASE WHEN (balls=3 AND pitch_result = 'B') or pitch_result = 'HBP' THEN 1 END) AS freebies "
    query += "FROM pitch_log_T "
    query += ending
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (freebies,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = freebies if freebies is not None else 0
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)
            
def insert_strikeout_count(cursora,new_sheetb,ending,row_i,col_i,exe):
    ###### Count of Strikeouts
    query = "SELECT COUNT(CASE WHEN strikes = 2 AND pitch_result IN ('SL','SS','SSC','D3SS') THEN 1 END) AS Ks "
    query += "FROM pitch_log_T "
    query += ending
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (Ks,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = Ks if Ks is not None else 0
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)
        
def insert_advantage_counts_percentage(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Percentage of Pitches Thrown in Advantage Counts
    query = "SELECT CASE WHEN COUNT(CASE WHEN pitch_result <>'0' THEN 1 END) > 0 "
    query+= "THEN (COUNT(CASE WHEN strikes>balls THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' THEN 1 END)) ELSE 0 END AS PTIAC FROM pitch_log_t "
    query+= ending
    cursora.execute(query, exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (ptiac,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = int(ptiac) if ptiac is not None else 0
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)     
        
def insert_disadvantage_counts_percentage(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Percentage of Pitches Thrown in DisAdvantage Counts
    query = "SELECT CASE WHEN COUNT(CASE WHEN pitch_result <>'0' THEN 1 END) > 0 "
    query+= "THEN (COUNT(CASE WHEN strikes<balls THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' THEN 1 END)) ELSE 0 END AS PTIDC FROM pitch_log_t "
    query+= ending
    cursora.execute(query, exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (ptidc,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = int(ptidc) if ptidc is not None else 0
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)

def insert_strikeout_percentage(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Strikeout Percentage
    query = "SELECT CASE WHEN COUNT(CASE WHEN ab_result <> '0' THEN 1 END) >0"
    query += " THEN (COUNT(CASE WHEN strikes = 2 AND (pitch_result = 'SL' or pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') THEN 1 END) * 100 /" 
    query += " COUNT(CASE WHEN ab_result <> '0' THEN 1 END)) ELSE 0 END AS KPer FROM pitch_log_t "
    query += ending
    cursora.execute(query, exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
    
    for k, (kper,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = int(kper) if kper is not None else 0
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)
        
def insert_ground_ball_out_percentage(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Ground Ball Out Percentage
    query = "SELECT CASE WHEN COUNT(CASE WHEN ab_result<>'0' THEN 1 END) >0"
    query += " THEN (COUNT(CASE WHEN bip_result = 'GO' OR bip_result = 'DP' THEN 1 END) * 100 /"
    query += " COUNT(CASE WHEN ab_result<>'0' THEN 1 END)) ELSE 0 END AS GBOPer FROM pitch_log_t "
    query += ending
    cursora.execute(query, exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
     
    for k, (gboper,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = int(gboper) if gboper is not None else 0
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)
        
def insert_fly_ball_out_percentage(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Fly Ball Out Percentage
    query = "SELECT CASE WHEN COUNT(CASE WHEN ab_result<>'0' THEN 1 END) >0"
    query += " THEN (COUNT(CASE WHEN bip_result = 'FO' THEN 1 END) * 100 /" 
    query += " COUNT(CASE WHEN ab_result<>'0' THEN 1 END)) ELSE 0 END AS FBOPer FROM pitch_log_t "
    query += ending
    cursora.execute(query, exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
     
    for k, (fboper,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in = int(fboper) if fboper is not None else 0
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)  
        
def insert_at_bat_win_rate(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Percentage of At Bats that resulted in an Out
    query = "SELECT CASE WHEN COUNT(CASE WHEN ab_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN ab_result = 'out' THEN 1 END) * 100.0 / "
    query += "COUNT(CASE WHEN ab_result <> '0' THEN 1 END)) ELSE 0 END AS WINRATE FROM pitch_log_t "
    query += ending
    
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
        
    for k,(winrate,) in enumerate(data,3):
        if (trip):
            row_i=k
        put_in=int(winrate) if winrate is not None else 0
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)
        
def insert_pitch_spread_percentage(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Fastball - Curveball - Slider - Change UP - Splitter Spread Percentage
    query = "SELECT "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'FF' AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'FF' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' THEN 1 END)) ELSE 0 END AS FFP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'CB' AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'CB' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' THEN 1 END)) ELSE 0 END AS CBP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'SL' AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'SL' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' THEN 1 END)) ELSE 0 END AS SLP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'CH' AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'CH' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' THEN 1 END)) ELSE 0 END AS CHP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'SP' AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'SP' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' THEN 1 END))ELSE 0 END AS SPP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'FT' AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'FT' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' THEN 1 END)) ELSE 0 END AS FTP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'CT' AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'CT' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' THEN 1 END))ELSE 0 END AS CTP "
    query += "FROM pitch_log_t "
    query +=ending
    
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
        
    for k,(FFP,CBP,SLP,CHP,SPP,FTP,CTP) in enumerate(data,3):
        if (trip):
            row_i=k
        FFP=int(FFP) if FFP is not None else 0
        CBP=int(CBP) if CBP is not None else 0
        SLP=int(SLP) if SLP is not None else 0
        CHP=int(CHP) if CHP is not None else 0
        SPP=int(SPP) if SPP is not None else 0
        FTP=int(FTP) if FTP is not None else 0
        CTP=int(CTP) if CTP is not None else 0  
        
        put_in = f"{FFP}-{CBP}-{SLP}-{CHP}-{SPP}-{FTP}-{CTP}"
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)
        
def insert_pitch_spread_strike_percentage(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Four Seam - Curveball - Slider - Change UP - Splitter - Two Seam - Cutter Spread Strike Percentage
    query = "SELECT "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'FF' AND pitch_result <> 'B' AND pitch_result <> 'HBP' AND pitch_result <> '0' THEN 1 END) > 0 " 
    query += "THEN (COUNT(CASE WHEN pitch_type = 'FF' AND pitch_result <> 'B' AND pitch_result <> 'HBP' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' AND pitch_type = 'FF' THEN 1 END)) ELSE 0 END AS FFP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'CB' AND pitch_result <> 'B' AND pitch_result <> 'HBP' AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'CB' AND pitch_result <> 'B' AND pitch_result <> 'HBP' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' AND pitch_type = 'CB' THEN 1 END)) ELSE 0 END AS CBP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'SL' AND pitch_result <> 'B' AND pitch_result <> 'HBP' AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'SL' AND pitch_result <> 'B' AND pitch_result <> 'HBP' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' AND pitch_type = 'SL' THEN 1 END)) ELSE 0 END AS SLP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'CH' AND pitch_result <> 'B' AND pitch_result <> 'HBP' AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'CH' AND pitch_result <> 'B' AND pitch_result <> 'HBP' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' AND pitch_type = 'CH' THEN 1 END)) ELSE 0 END AS CHP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'SP' AND pitch_result <> 'B' AND pitch_result <> 'HBP' AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'SP' AND pitch_result <> 'B' AND pitch_result <> 'HBP' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' AND pitch_type = 'SP' THEN 1 END)) ELSE 0 END AS SPP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'FT' AND pitch_result <> 'B' AND pitch_result <> 'HBP' AND pitch_result <> '0' THEN 1 END) > 0 " 
    query += "THEN (COUNT(CASE WHEN pitch_type = 'FT' AND pitch_result <> 'B' AND pitch_result <> 'HBP' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' AND pitch_type = 'FT' THEN 1 END)) ELSE 0 END AS FTP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'CT' AND pitch_result <> 'B' AND pitch_result <> 'HBP' AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'CT' AND pitch_result <> 'B' AND pitch_result <> 'HBP' THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result <> '0' AND pitch_type = 'CT' THEN 1 END)) ELSE 0 END AS CTP "
    query += "FROM pitch_log_t "
    query +=ending
    
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
        
    for k,(FFP,CBP,SLP,CHP,SPP,FTP,CTP) in enumerate(data,3):
        if (trip):
            row_i=k
        FFP=int(FFP) if FFP is not None else 0
        CBP=int(CBP) if CBP is not None else 0
        SLP=int(SLP) if SLP is not None else 0
        CHP=int(CHP) if CHP is not None else 0
        SPP=int(SPP) if SPP is not None else 0
        FTP=int(FTP) if FTP is not None else 0 
        CTP=int(CTP) if CTP is not None else 0  
                
        put_in = f"{FFP}-{CBP}-{SLP}-{CHP}-{SPP}-{FTP}-{CTP}"
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)       
        
def insert_pitch_spread_whiff_percentage(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Fastball - Curveball - Slider - Change UP - Splitter Spread whiff Percentage
    query = "SELECT "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'FF' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'FF' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result IN('SS','SSC','D3SS','F','BIP') AND pitch_type = 'FF' THEN 1 END)) ELSE 0 END AS FFP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'CB' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'CB' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result IN('SS','SSC','D3SS','F','BIP') AND pitch_type = 'CB' THEN 1 END)) ELSE 0 END AS CBP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'SL' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'SL' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result IN('SS','SSC','D3SS','F','BIP') AND pitch_type = 'SL' THEN 1 END)) ELSE 0 END AS SLP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'CH' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'CH' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result IN('SS','SSC','D3SS','F','BIP') AND pitch_type = 'CH' THEN 1 END)) ELSE 0 END AS CHP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'SP' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS')  AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'SP' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result IN('SS','SSC','D3SS','F','BIP') AND pitch_type = 'SP' THEN 1 END)) ELSE 0 END AS SPP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'FT' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'FT' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result IN('SS','SSC','D3SS','F','BIP') AND pitch_type = 'FT' THEN 1 END)) ELSE 0 END AS FTP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'CT' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'CT' AND (pitch_result = 'SS' or pitch_result = 'SSC' or pitch_result = 'D3SS') THEN 1 END) * 100.0 / COUNT(CASE WHEN pitch_result IN('SS','SSC','D3SS','F','BIP') AND pitch_type = 'CT' THEN 1 END)) ELSE 0 END AS CTP "
    query += "FROM pitch_log_t "
    query +=ending
    
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
        
    for k,(FFP,CBP,SLP,CHP,SPP,FTP,CTP) in enumerate(data,3):
        if (trip):
            row_i=k
        FFP=int(FFP) if FFP is not None else 0
        CBP=int(CBP) if CBP is not None else 0
        SLP=int(SLP) if SLP is not None else 0
        CHP=int(CHP) if CHP is not None else 0
        SPP=int(SPP) if SPP is not None else 0
        FTP=int(FTP) if FTP is not None else 0 
        CTP=int(CTP) if CTP is not None else 0  
                
        put_in = f"{FFP}-{CBP}-{SLP}-{CHP}-{SPP}-{FTP}-{CTP}"
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)
        
def insert_pitch_spread_hits_percentage(cursora,new_sheetb,ending,row_i,col_i,exe):
    ##### Fastball - Curveball - Slider - Change UP - Splitter Spread Hit Percentage
    query = "SELECT "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'FF' AND bip_result in ('1B','2B','3B','HR') THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'FF' AND bip_result in ('1B','2B','3B','HR') THEN 1 END) * 100.0 / COUNT(CASE WHEN bip_result in ('1B','2B','3B','HR') THEN 1 END)) ELSE 0 END AS FFP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'CB' AND bip_result in ('1B','2B','3B','HR') THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'CB' AND bip_result in ('1B','2B','3B','HR') THEN 1 END) * 100.0 / COUNT(CASE WHEN bip_result in ('1B','2B','3B','HR') THEN 1 END)) ELSE 0 END AS CBP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'SL' AND bip_result in ('1B','2B','3B','HR') AND pitch_result <> '0' THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'SL' AND bip_result in ('1B','2B','3B','HR') THEN 1 END) * 100.0 / COUNT(CASE WHEN bip_result in ('1B','2B','3B','HR') THEN 1 END)) ELSE 0 END AS SLP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'CH' AND bip_result in ('1B','2B','3B','HR') THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'CH' AND bip_result in ('1B','2B','3B','HR') THEN 1 END) * 100.0 / COUNT(CASE WHEN bip_result in ('1B','2B','3B','HR') THEN 1 END)) ELSE 0 END AS CHP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'SP' AND bip_result in ('1B','2B','3B','HR') THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'SP' AND bip_result in ('1B','2B','3B','HR') THEN 1 END) * 100.0 / COUNT(CASE WHEN bip_result in ('1B','2B','3B','HR') THEN 1 END)) ELSE 0 END AS SPP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'FT' AND bip_result in ('1B','2B','3B','HR') THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'FT' AND bip_result in ('1B','2B','3B','HR') THEN 1 END) * 100.0 / COUNT(CASE WHEN bip_result in ('1B','2B','3B','HR') THEN 1 END)) ELSE 0 END AS FTP, "
    query += "CASE WHEN COUNT(CASE WHEN pitch_type = 'CT' AND bip_result in ('1B','2B','3B','HR') THEN 1 END) > 0 "
    query += "THEN (COUNT(CASE WHEN pitch_type = 'CT' AND bip_result in ('1B','2B','3B','HR') THEN 1 END) * 100.0 / COUNT(CASE WHEN bip_result in ('1B','2B','3B','HR') THEN 1 END)) ELSE 0 END AS CUP "
    query += "FROM pitch_log_t "
    query +=ending
    
    cursora.execute(query,exe)
    data=cursora.fetchall()
    trip=False
    if (row_i==0):
        trip=True
        
    for k,(FFP,CBP,SLP,CHP,SPP,FTP,CTP) in enumerate(data,3):
        if (trip):
            row_i=k
        FFP=int(FFP) if FFP is not None else 0
        CBP=int(CBP) if CBP is not None else 0
        SLP=int(SLP) if SLP is not None else 0
        CHP=int(CHP) if CHP is not None else 0
        SPP=int(SPP) if SPP is not None else 0
        FTP=int(FTP) if FTP is not None else 0 
        CTP=int(CTP) if CTP is not None else 0  
                
        put_in = f"{FFP}-{CBP}-{SLP}-{CHP}-{SPP}-{FTP}-{CTP}"
        new_sheetb.cell(row=row_i, column=col_i, value=put_in)
               
def insert_what_got_hit(cursora,new_sheetb,start_row,exe):
    
    insert_secondary_header(start_row-2,1,"PITCHES THAT GOT HIT",new_sheetb)
    insert_secondary_header(start_row-1,1,"Name",new_sheetb)
    insert_secondary_header(start_row-1,2,"Pitch Type",new_sheetb)
    insert_secondary_header(start_row-1,3,"Velo",new_sheetb)
    insert_secondary_header(start_row-1,4,"Count",new_sheetb)
    insert_secondary_header(start_row-1,5,"BIP Result",new_sheetb)
    insert_secondary_header(start_row-1,6,"Batter Number",new_sheetb)
    
    
    query = "SELECT fname, lname, pitch_type,velocity,balls,strikes,bip_result,batter_number "
    query += "FROM pitch_log_T WHERE bip_result in ('1B','2B','3B','HR') and date = %s and opponent = %s"
    query += "Order By pitch_id ASC"
    cursora.execute(query,exe)
    data=cursora.fetchall()
    
    for k,(fname,lname,pitch_type,velocity,balls,strikes,bip_result,batter_number) in enumerate(data,start_row):
        put_in=f"{fname} {lname}"
        new_sheetb.cell(row=k,column=1,value=put_in)
        
        put_in= pitch_type
        new_sheetb.cell(row=k,column=2,value=put_in)
        
        put_in= int(velocity)
        new_sheetb.cell(row=k,column=3,value=put_in)
        
        put_in= f"{balls} - {strikes}"
        new_sheetb.cell(row=k,column=4,value=put_in)
        
        put_in= f"{bip_result}"
        new_sheetb.cell(row=k,column=5,value=put_in)
        
        put_in= f"#{batter_number}"
        new_sheetb.cell(row=k,column=6,value=put_in)
        
def insert_whip_by_inning_of_work(cursora,new_sheetb,firstname,lastname,start_row):
    ##### tracks whip as a pitcher goes deeper into games
    
    insert_secondary_header(start_row-2,1,"WHIP IN EACH INNING OF WORK",new_sheetb)
    insert_secondary_header(start_row-1,1,"First",new_sheetb)
    insert_secondary_header(start_row-1,2,"Second",new_sheetb)
    insert_secondary_header(start_row-1,3,"Third",new_sheetb)
    insert_secondary_header(start_row-1,4,"Fourth",new_sheetb)
    insert_secondary_header(start_row-1,5,"Fifth",new_sheetb)
    insert_secondary_header(start_row-1,6,"Sixth",new_sheetb)
    insert_secondary_header(start_row-1,7,"Seventh",new_sheetb)
    insert_secondary_header(start_row-1,8,"Eight",new_sheetb)
    insert_secondary_header(start_row-1,9,"Ninth",new_sheetb)
    
    query = "SELECT "
    query += "COUNT(CASE WHEN ab_result='out' AND outs_accrued IN (1,2,3) THEN 1 END) AS FirstOuts, "
    query += "COUNT(CASE WHEN ab_result='safe' AND bip_result NOT IN ('E','HBP') AND outs_accrued IN (0,1,2) THEN 1 END) AS FirstSafes, "
    
    query += "COUNT(CASE WHEN ab_result='out' AND outs_accrued IN (4,5,6) THEN 1 END) AS SecondOuts, "
    query += "COUNT(CASE WHEN ab_result='safe' AND bip_result NOT IN ('E','HBP') AND outs_accrued IN (3,4,5) THEN 1 END) AS SecondSafes, "
    
    query += "COUNT(CASE WHEN ab_result='out' AND outs_accrued IN (7,8,9) THEN 1 END) AS ThirdOuts, "
    query += "COUNT(CASE WHEN ab_result='safe' AND bip_result NOT IN ('E','HBP') AND outs_accrued IN (6,7,8) THEN 1 END) AS ThirdSafes, "
    
    query += "COUNT(CASE WHEN ab_result='out' AND outs_accrued IN (10,11,12) THEN 1 END) AS FourthOuts, "
    query += "COUNT(CASE WHEN ab_result='safe' AND bip_result NOT IN ('E','HBP') AND outs_accrued IN (9,10,11) THEN 1 END) AS FourthSafes, "
    
    query += "COUNT(CASE WHEN ab_result='out' AND outs_accrued IN (13,14,15) THEN 1 END) AS FifthOuts, "
    query += "COUNT(CASE WHEN ab_result='safe' AND bip_result NOT IN ('E','HBP') AND outs_accrued IN (12,13,14) THEN 1 END) AS FifthSafes, "
    
    query += "COUNT(CASE WHEN ab_result='out' AND outs_accrued IN (16,17,18) THEN 1 END) AS SixthOuts, "
    query += "COUNT(CASE WHEN ab_result='safe' AND bip_result NOT IN ('E','HBP') AND outs_accrued IN (15,16,17) THEN 1 END) AS SixthSafes, "
    
    query += "COUNT(CASE WHEN ab_result='out' AND outs_accrued IN (19,20,21) THEN 1 END) AS SeventhOuts, "
    query += "COUNT(CASE WHEN ab_result='safe' AND bip_result NOT IN ('E','HBP') AND outs_accrued IN (18,19,20) THEN 1 END) AS SeventhSafes, "
    
    query += "COUNT(CASE WHEN ab_result='out' AND outs_accrued IN (21,22,23) THEN 1 END) AS EighthOuts, "
    query += "COUNT(CASE WHEN ab_result='safe' AND bip_result NOT IN ('E','HBP') AND outs_accrued IN (21,22,23) THEN 1 END) AS EighthSafes, "
    
    query += "COUNT(CASE WHEN ab_result='out' AND outs_accrued IN (24,25,26) THEN 1 END) AS NinthOuts, "
    query += "COUNT(CASE WHEN ab_result='safe' AND bip_result NOT IN ('E','HBP') AND outs_accrued IN (24,25,26) THEN 1 END) AS NinthSafes "
    
    query += "FROM pitch_log_t WHERE fname=%s AND lname=%s AND opponent <> 'Scrimmage'"
        
    cursora.execute(query,(firstname,lastname))
    data=cursora.fetchall()
    
    for k,(FirstOuts,FirstSafes,SecondOuts,SecondSafes,ThirdOuts,ThirdSafes,FourthOuts,FourthSafes,FifthOuts,FifthSafes,SixthOuts,SixthSafes,SeventhOuts,SeventhSafes,EighthOuts,EighthSafes,NinthOuts,NinthSafes) in enumerate(data,start_row):
        
        First=round(FirstSafes/(FirstOuts/3),3) if FirstOuts!=0 else "NA"
        Second=round(SecondSafes/(SecondOuts/3),3) if SecondOuts!=0 else "NA"
        Third=round(ThirdSafes/(ThirdOuts/3),3) if ThirdOuts!=0 else "NA"
        Fourth=round(FourthSafes/(FourthOuts/3),3) if FourthOuts!=0 else "NA"
        Fifth=round(FifthSafes/(FifthOuts/3),3) if FifthOuts!=0 else "NA"
        Sixth=round(SixthSafes/(SixthOuts/3),3) if SixthOuts!=0 else "NA"
        Seventh=round(SeventhSafes/(SeventhOuts/3),3) if SeventhOuts!=0 else "NA"
        Eighth=round(EighthSafes/(EighthOuts/3),3) if EighthOuts!=0 else "NA"
        Ninth=round(NinthSafes/(NinthOuts/3),3) if NinthOuts!=0 else "NA"
        
        
        new_sheetb.cell(row=k,column=1,value=First)
        new_sheetb.cell(row=k,column=2,value=Second)
        new_sheetb.cell(row=k,column=3,value=Third)
        new_sheetb.cell(row=k,column=4,value=Fourth)
        new_sheetb.cell(row=k,column=5,value=Fifth)
        new_sheetb.cell(row=k,column=6,value=Sixth)
        new_sheetb.cell(row=k,column=7,value=Seventh)
        new_sheetb.cell(row=k,column=8,value=Eighth)
        new_sheetb.cell(row=k,column=9,value=Ninth)
        

def insert_avg_peak_FB_velo_over_time_chart(cursora,new_sheetb,firstname,lastname,start_row):
    
    query = "SELECT date_n AS time, MAX(velocity) AS maxi, AVG(velocity) AS average "
    query+= "FROM pitch_log_T WHERE pitch_type IN ('FF','FT') AND fname=%s AND lname=%s AND pitch_count>0 GROUP BY date_n"
    
    cursora.execute(query,(firstname,lastname))
    data=cursora.fetchall()
    
    new_sheetb.cell(row=start_row,column=1,value="Time")
    new_sheetb.cell(row=start_row,column=2,value="Max Velo")
    new_sheetb.cell(row=start_row,column=3,value="Average Velo")
   
    
    for k,(time,maxi,average) in enumerate(data, start_row+1):
        new_sheetb.cell(row=k,column=1,value=time)
        new_sheetb.cell(row=k,column=2,value=maxi)
        new_sheetb.cell(row=k,column=3,value=average)
        
        
    chart = ScatterChart()
    chart.title = "Fastball Velocity Peaks and Averages Over Time"
    chart.x_axis.title = "Date"
    chart.y_axis.title = "Velocity (mph)"

    xvalues = Reference(new_sheetb, min_col=1, min_row=start_row+1, max_row=start_row+len(data))
    yvalues1 = Reference(new_sheetb, min_col=2, min_row=start_row+1, max_row=start_row+len(data))
    yvalues2 = Reference(new_sheetb, min_col=3,min_row=start_row+1, max_row=start_row+len(data))
    
    series1 = Series(yvalues1, xvalues, title="Max Velo")
    series2 = Series(yvalues2, xvalues, title="Avg Velo")
    
    chart.series.append(series1)
    chart.series.append(series2)

    new_sheetb.add_chart(chart, "A{}".format(start_row))   
    
def insert_movement_profile_chart(cursora,new_sheetb,firstname,lastname,start_row):
    query = "SELECT inducedvertbreak AS Vert, horzbreak AS Horz, pitch_type "
    query += "FROM trackman_pitching_data_t WHERE pfname = %s and plname=%s"
    
    cursora.execute(query,(firstname,lastname))
    data=cursora.fetchall()
    
    # Create a scatter plot using matplotlib
    plt.figure(figsize=(8, 6))
    plt.title("Movement Profiles")
    plt.xlabel("Horizontal Break")
    plt.ylabel("Induced Vertical Break")
    plt.grid(True)
    
    plt.xlim(-30, 30)
    plt.ylim(-30, 30)
    
    # Define colors for different pitch types
    colors = {'Fastball': 'r','Four-Seam':'r', 'Curveball': 'g', 'ChangeUp': 'b', 'Slider': 'c', 'TwoSeamFastBall': 'm', 'Sinker':'y','Cutter':'#FFA500'}

    scatter_objs = []

    for pitch_type, color in colors.items():
    # Filter data for current pitch type
        filtered_data = [row for row in data if row[2] == pitch_type]
        if filtered_data:
        # Extract x and y values for scatter plot
            x_values = [row[1] for row in filtered_data]
            y_values = [row[0] for row in filtered_data]
            # Plot scatter plot for current pitch type
            scatter_obj = plt.scatter(x_values, y_values, color=color, label=pitch_type)
            scatter_objs.append(scatter_obj)
        
    if len(data)!=0:
        plt.legend(loc='lower right')
        
    plt.axhline(0, color='k', linewidth=1.5)  # y=0 line
    plt.axvline(0, color='k', linewidth=1.5)  # x=0 line

    # Save the plot to a BytesIO object
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)

    # Insert the image into the Excel file
    img = Image(buffer)
    new_sheetb.add_image(img, "F{}".format(start_row))
    
    plt.close()

    
def insert_avg_pitch_velo_over_time(cursor, new_sheet, firstname, lastname, start_row):
    pitch_types = ['FF', 'CB', 'SL', 'CH', 'SP', 'FT', 'CT']
    chart = ScatterChart()
    chart.title = "Average Pitch Velocity Over Time"
    chart.x_axis.title = "Date"
    chart.y_axis.title = "Velocity (mph)"
    
    # Starting column index
    col_index = 13
    
    for pitch_type in pitch_types:
        query = "SELECT date_n AS time, AVG(velocity) as AVG FROM pitch_log_T WHERE pitch_type=%s AND fname=%s AND lname=%s AND pitch_count>0 GROUP BY date_n"
        cursor.execute(query, (pitch_type, firstname, lastname))
        data = cursor.fetchall()
        
        # Insert data into spreadsheet
        new_sheet.cell(row=start_row, column=col_index, value=f"{pitch_type} Time")
        new_sheet.cell(row=start_row, column=col_index + 1, value=f"{pitch_type} Velo")
        
        for k, (time, average) in enumerate(data, start=start_row+1):
            new_sheet.cell(row=k, column=col_index, value=time)
            new_sheet.cell(row=k, column=col_index + 1, value=average)
        
        # Add data to chart
        if (len(data)!=0):
            x_values = Reference(new_sheet, min_col=col_index, min_row=start_row+1, max_row=start_row+len(data))
            y_values = Reference(new_sheet, min_col=col_index + 1, min_row=start_row+1, max_row=start_row+len(data))
            series = Series(y_values, x_values, title=pitch_type)
            chart.series.append(series)
        
        # Increment column index for the next pitch type
        col_index += 2
        
    query = "SELECT MIN(velocity),MAX(velocity) FROM pitch_log_T WHERE fname=%s AND lname=%s AND pitch_count>0"
    cursor.execute(query,(firstname,lastname))
    data=cursor.fetchone()
    
    chart.y_axis.scaling.min = data[0]-2  # Set minimum value for y-axis
    chart.y_axis.scaling.max = data[1]+2  # Set maximum value for y-axis
    
    new_sheet.add_chart(chart, "M{}".format(start_row))
    
    
def up_pitchers_log(cursor,update_date,file_name):
           
           workbook=create_workbook(file_name)[0]
           file_path=create_workbook(file_name)[1]
           
           sheets=[]
           for sheet in workbook.worksheets:
               if sheet.title!='Sheet':
                   sheets.append(sheet.title)
                   
           names=[]
           new_name_query="SELECT DISTINCT fname,lname FROM pitch_log_T WHERE pitch_id <>'0' "
           cursor.execute(new_name_query,())
           data=cursor.fetchall()
           
           for row in data:
               fname,lname=row
               names.append(fname + " " + lname)
               
           for sheet in sheets:
               if sheet in names:
                   names.remove(sheet)
                     
           names_check = names
           
           if (len(names_check)!=0):
               for iname in names_check:
                   new_sheet=setup(iname,workbook,"Name",iname,"Updated Date",update_date,"","",header_pos,pitcher_headers)
                   firstname,lastname=iname.split(" ")
                   
                   query="SELECT DISTINCT date AS datea, date_n FROM pitch_log_t WHERE fname=%s AND lname = %s and pitch_id <> '0' ORDER BY date_n ASC"
                   cursor.execute(query,(firstname,lastname))
                   data=cursor.fetchall()
                   
                   total_pitch_count=0
                   total_innings=0
                   total_peak_velo=0
                   pitchers=0
                   
                   
                   for j, (datea) in enumerate(data,3):
                       dates=len(data)
                       datea =datea[0] if isinstance(datea, tuple) and datea else datea
                       new_sheet.cell(row=j, column=1, value=datea)
                       
                       exea=(datea,firstname,lastname)
                       iplEndStatement="WHERE date = %s AND fname = %s AND lname = %s AND pitch_id <> '0' "
                               
                       insert_oppo(cursor, new_sheet,iplEndStatement,j,2,exea)
                       
                       total_pitch_count+=insert_pitches_thrown(cursor, new_sheet, iplEndStatement, j, 3, exea)
                       
                       total_innings+=insert_pitches_per_inning(cursor, new_sheet, iplEndStatement, j, 4, exea,False)
                       
                       a, b=insert_peak_velo(cursor, new_sheet, iplEndStatement, j, 5, exea)
                      
                       total_peak_velo+=a
                       pitchers+=b
                                   
                       insert_1st_pitch_strike_percentage(cursor, new_sheet, iplEndStatement, j, 6, exea, 60)
                       
                       insert_off_speed_strike_percentage(cursor, new_sheet, iplEndStatement, j, 7, exea, 50)
                     
                       insert_swing_and_miss_percentage(cursor, new_sheet,iplEndStatement, j, 8, exea, 25)
                            
                       insert_velo_range(cursor, new_sheet, iplEndStatement, j, 9, exea,False,False,False)        
                               
                       insert_chases(cursor, new_sheet, iplEndStatement, j, 10, exea, False, 0)
                       
                       insert_ahead_after_3_pitches_percentage(cursor,new_sheet,iplEndStatement,j,11,exea,60)
                       
                       insert_opponent_slugging_percentage(cursor, new_sheet, iplEndStatement, j, 12, exea, .4)
                       
                       insert_WHIP(cursor, new_sheet, iplEndStatement, j, 13, exea, False,False, 0)
                       
                       insert_OBP(cursor, new_sheet, iplEndStatement, j, 14, exea)
                       
                       insert_OPS(cursor, new_sheet, iplEndStatement, j, 15, exea)
                       
                       insert_lead_off_out_percentage(cursor, new_sheet, iplEndStatement, j, 16, exea, 65)
                       
                       insert_overall_strike_percentage(cursor, new_sheet, iplEndStatement, j, 17, exea, 60)
                       
                       insert_baa(cursor,new_sheet,iplEndStatement,j,18,exea)
                       
                       insert_baa_with_2_strikes(cursor, new_sheet, iplEndStatement, j, 19, exea, 0.15)
                       
                       insert_baa_bip(cursor, new_sheet, iplEndStatement, j, 20, exea)
                       
                       insert_freebases_count(cursor, new_sheet, iplEndStatement, j, 21, exea)
                       
                       insert_strikeout_count(cursor, new_sheet, iplEndStatement, j, 22, exea)
                               
                       insert_advantage_counts_percentage(cursor, new_sheet, iplEndStatement, j, 23, exea)
                       
                       insert_disadvantage_counts_percentage(cursor, new_sheet, iplEndStatement, j, 24, exea)
                                          
                       insert_strikeout_percentage(cursor, new_sheet, iplEndStatement, j, 25, exea)
                       
                       insert_ground_ball_out_percentage(cursor, new_sheet, iplEndStatement, j, 26, exea)
                       
                       insert_fly_ball_out_percentage(cursor, new_sheet, iplEndStatement, j, 27, exea)
                       
                       insert_at_bat_win_rate(cursor, new_sheet, iplEndStatement, j, 28, exea)
                       
                       insert_pitch_spread_percentage(cursor, new_sheet, iplEndStatement, j, 29, exea)
                       
                       insert_pitch_spread_strike_percentage(cursor, new_sheet, iplEndStatement, j, 30, exea)
                       
                       insert_pitch_spread_whiff_percentage(cursor, new_sheet, iplEndStatement, j, 31, exea)
                       
                       insert_pitch_spread_hits_percentage(cursor, new_sheet, iplEndStatement, j, 32, exea)
                       
                   ######################## Player's Season Totals ####################################
                   
                   iplEndStatement="WHERE fname = %s AND lname = %s AND pitch_id <> '0' AND opponent <> 'Scrimmage' "
                   exea=(firstname,lastname)
                   
                   # Calculate and insert team totals
                   season_totals_row = dates + 4  # Assuming a gap of one row between individual pitchers and team totals
                   
                   new_sheet.cell(row=season_totals_row, column=1, value="Season Totals")
                
                   ##### Total Pitch Count
                   new_sheet.cell(row=season_totals_row, column=2, value="-")
            
                   ##### Total Pitch Count
                   new_sheet.cell(row=season_totals_row, column=3, value=total_pitch_count)
            
                   ##### Average Pitches Per Inning
                   if total_innings !=0:
                       new_sheet.cell(row=season_totals_row, column=4, value=round(total_pitch_count/total_innings,2))
                   else:
                       new_sheet.cell(row=season_totals_row,column = 4, value = 0)
                
                   #####  Average Peak Velocity
                   if pitchers!=0:
                       new_sheet.cell(row=season_totals_row, column=5, value=round(total_peak_velo/pitchers,2))
                   else:
                       new_sheet.cell(row=season_totals_row,column = 5, value = 0)
                   
                   
                   insert_1st_pitch_strike_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 6, exea, 60)
                   
                   insert_off_speed_strike_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 7, exea, 50)
                 
                   insert_swing_and_miss_percentage(cursor, new_sheet,iplEndStatement, season_totals_row, 8, exea, 25)
                        
                   insert_velo_range(cursor, new_sheet, iplEndStatement, season_totals_row, 9, exea,False,False,False)        
                           
                   insert_chases(cursor, new_sheet, iplEndStatement, season_totals_row, 10, exea, False, total_innings)
                   
                   insert_ahead_after_3_pitches_percentage(cursor,new_sheet,iplEndStatement,season_totals_row,11,exea,60)
                   
                   insert_opponent_slugging_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 12, exea, .4)
                   
                   insert_WHIP(cursor, new_sheet, iplEndStatement, season_totals_row, 13, exea, False,False, total_innings)
                   
                   insert_OBP(cursor, new_sheet, iplEndStatement, season_totals_row, 14, exea)
                   
                   insert_OPS(cursor, new_sheet, iplEndStatement, season_totals_row, 15, exea)
                   
                   insert_lead_off_out_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 16, exea, 65)
                   
                   insert_overall_strike_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 17, exea, 60)
                   
                   insert_baa(cursor,new_sheet,iplEndStatement,season_totals_row,18,exea)
                   
                   insert_baa_with_2_strikes(cursor, new_sheet, iplEndStatement, season_totals_row, 19, exea, 0.15)
                   
                   insert_baa_bip(cursor, new_sheet, iplEndStatement, season_totals_row, 20, exea)
                   
                   insert_freebases_count(cursor, new_sheet, iplEndStatement, season_totals_row, 21, exea)
                   
                   insert_strikeout_count(cursor, new_sheet, iplEndStatement, season_totals_row, 22, exea)
                   
                   insert_advantage_counts_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 23, exea)
                   
                   insert_disadvantage_counts_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 24, exea)
                                      
                   insert_strikeout_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 25, exea)
                   
                   insert_ground_ball_out_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 26, exea)
                   
                   insert_fly_ball_out_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 27, exea)
                   
                   insert_at_bat_win_rate(cursor, new_sheet, iplEndStatement, season_totals_row, 28, exea)
                   
                   insert_pitch_spread_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 29, exea)
                   
                   insert_pitch_spread_strike_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 30, exea)
                   
                   insert_pitch_spread_whiff_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 31, exea)  
                   
                   insert_pitch_spread_hits_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 32, exea)
                                      
                   bold_first_column_if_threshold(new_sheet, 5)
                   
                   insert_whip_by_inning_of_work(cursor, new_sheet, firstname, lastname, season_totals_row+4)
                   
                   insert_avg_peak_FB_velo_over_time_chart(cursor, new_sheet, firstname, lastname, season_totals_row+6)
                   
                   insert_movement_profile_chart(cursor, new_sheet, firstname, lastname, season_totals_row+6)
                   
                   insert_avg_pitch_velo_over_time(cursor, new_sheet, firstname, lastname, season_totals_row+6)
                   
                   adjust_formating(new_sheet, season_totals_row)
                   
           for sheet in workbook.worksheets:
               if sheet.title!='Sheet':
                   
                   fullname=sheet.title.split(" ")
                   
                   firstname=fullname[0]
                   lastname=fullname[1]
                   
                   
                   
                   query="SELECT MAX(date_n) FROM pitch_log_T WHERE fname=%s AND lname=%s"
                   cursor.execute(query,(firstname,lastname))
                   data=cursor.fetchone()
                   
                   last_pitched=data[0]
                   last_pitched=str(last_pitched.strftime("%b %d"))
                   last_pitched=last_pitched.split(" ")
                   last_pitched_month=last_pitched[0]
                   last_pitched_day=str(int(last_pitched[1]))
                   last_pitched=last_pitched_month + " " + last_pitched_day
                   
                   for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
                       cell = row[0]
                       if cell.value is None:
                           empty_row = cell.row
                           break
                   
                   last_updated = str(sheet.cell(row=empty_row - 1, column=1).value)
                   month, day, year = map(int, last_updated.split("-"))

                   months = [
                           "Jan", "Feb", "Mar", "Apr", "May", "Jun", 
                           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
                           ]

                   last_updated = f"{months[month-1]} {day}"
                   
                   
                   if (last_pitched!=last_updated):
                       del workbook[sheet.title]
                       new_sheet=setup(sheet.title,workbook,"Name",sheet.title,"Updated Date",update_date,"","",header_pos,pitcher_headers)
                       
                       query="SELECT DISTINCT date AS datea, date_n FROM pitch_log_t WHERE fname=%s AND lname = %s and pitch_id <> '0' ORDER BY date_n ASC"
                       cursor.execute(query,(firstname,lastname))
                       data=cursor.fetchall()
                       
                       total_pitch_count=0
                       total_innings=0
                       total_peak_velo=0
                       pitchers=0
                       
                       
                       for j, (datea) in enumerate(data,3):
                           dates=len(data)
                           datea =datea[0] if isinstance(datea, tuple) and datea else datea
                           new_sheet.cell(row=j, column=1, value=datea)
                           
                           exea=(datea,firstname,lastname)
                           iplEndStatement="WHERE date = %s AND fname = %s AND lname = %s AND pitch_id <> '0' "
                                   
                           insert_oppo(cursor, new_sheet,iplEndStatement,j,2,exea)
                           
                           total_pitch_count+=insert_pitches_thrown(cursor, new_sheet, iplEndStatement, j, 3, exea)
                           
                           total_innings+=insert_pitches_per_inning(cursor, new_sheet, iplEndStatement, j, 4, exea,False)
                           
                           a, b=insert_peak_velo(cursor, new_sheet, iplEndStatement, j, 5, exea)
                          
                           total_peak_velo+=a
                           pitchers+=b
                                       
                           insert_1st_pitch_strike_percentage(cursor, new_sheet, iplEndStatement, j, 6, exea, 60)
                           
                           insert_off_speed_strike_percentage(cursor, new_sheet, iplEndStatement, j, 7, exea, 50)
                         
                           insert_swing_and_miss_percentage(cursor, new_sheet,iplEndStatement, j, 8, exea, 25)
                                
                           insert_velo_range(cursor, new_sheet, iplEndStatement, j, 9, exea,False,False,False)        
                                   
                           insert_chases(cursor, new_sheet, iplEndStatement, j, 10, exea, False, 0)
                           
                           insert_ahead_after_3_pitches_percentage(cursor,new_sheet,iplEndStatement,j,11,exea,60)
                           
                           insert_opponent_slugging_percentage(cursor, new_sheet, iplEndStatement, j, 12, exea, .4)
                           
                           insert_WHIP(cursor, new_sheet, iplEndStatement, j, 13, exea, False,False, 0)
                           
                           insert_OBP(cursor, new_sheet, iplEndStatement, j, 14, exea)
                           
                           insert_OPS(cursor, new_sheet, iplEndStatement, j, 15, exea)
                           
                           insert_lead_off_out_percentage(cursor, new_sheet, iplEndStatement, j, 16, exea, 65)
                           
                           insert_overall_strike_percentage(cursor, new_sheet, iplEndStatement, j, 17, exea, 60)
                           
                           insert_baa(cursor,new_sheet,iplEndStatement,j,18,exea)
                           
                           insert_baa_with_2_strikes(cursor, new_sheet, iplEndStatement, j, 19, exea, 0.15)
                           
                           insert_baa_bip(cursor, new_sheet, iplEndStatement, j, 20, exea)
                           
                           insert_freebases_count(cursor, new_sheet, iplEndStatement, j, 21, exea)
                           
                           insert_strikeout_count(cursor, new_sheet, iplEndStatement, j, 22, exea)
                                   
                           insert_advantage_counts_percentage(cursor, new_sheet, iplEndStatement, j, 23, exea)
                           
                           insert_disadvantage_counts_percentage(cursor, new_sheet, iplEndStatement, j, 24, exea)
                                              
                           insert_strikeout_percentage(cursor, new_sheet, iplEndStatement, j, 25, exea)
                           
                           insert_ground_ball_out_percentage(cursor, new_sheet, iplEndStatement, j, 26, exea)
                           
                           insert_fly_ball_out_percentage(cursor, new_sheet, iplEndStatement, j, 27, exea)
                           
                           insert_at_bat_win_rate(cursor, new_sheet, iplEndStatement, j, 28, exea)
                           
                           insert_pitch_spread_percentage(cursor, new_sheet, iplEndStatement, j, 29, exea)
                           
                           insert_pitch_spread_strike_percentage(cursor, new_sheet, iplEndStatement, j, 30, exea)
                           
                           insert_pitch_spread_whiff_percentage(cursor, new_sheet, iplEndStatement, j, 31, exea)
                           
                           insert_pitch_spread_hits_percentage(cursor, new_sheet, iplEndStatement, j, 32, exea)
                           
                       ######################## Player's Season Totals ####################################
                       
                       iplEndStatement="WHERE fname = %s AND lname = %s AND pitch_id <> '0' AND opponent <> 'Scrimmage' "
                       exea=(firstname,lastname)
                       
                       # Calculate and insert team totals
                       season_totals_row = dates + 4  # Assuming a gap of one row between individual pitchers and team totals
                       
                       new_sheet.cell(row=season_totals_row, column=1, value="Season Totals")
                    
                       ##### Total Pitch Count
                       new_sheet.cell(row=season_totals_row, column=2, value="-")
                
                       ##### Total Pitch Count
                       new_sheet.cell(row=season_totals_row, column=3, value=total_pitch_count)
                
                       ##### Average Pitches Per Inning
                       if total_innings !=0:
                           new_sheet.cell(row=season_totals_row, column=4, value=round(total_pitch_count/total_innings,2))
                       else:
                           new_sheet.cell(row=season_totals_row,column = 4, value = 0)
                    
                       #####  Average Peak Velocity
                       if pitchers!=0:
                           new_sheet.cell(row=season_totals_row, column=5, value=round(total_peak_velo/pitchers,2))
                       else:
                           new_sheet.cell(row=season_totals_row,column = 5, value = 0)
                       
                       
                       insert_1st_pitch_strike_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 6, exea, 60)
                       
                       insert_off_speed_strike_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 7, exea, 50)
                     
                       insert_swing_and_miss_percentage(cursor, new_sheet,iplEndStatement, season_totals_row, 8, exea, 25)
                            
                       insert_velo_range(cursor, new_sheet, iplEndStatement, season_totals_row, 9, exea,False,False,False)        
                               
                       insert_chases(cursor, new_sheet, iplEndStatement, season_totals_row, 10, exea, False, total_innings)
                       
                       insert_ahead_after_3_pitches_percentage(cursor,new_sheet,iplEndStatement,season_totals_row,11,exea,60)
                       
                       insert_opponent_slugging_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 12, exea, .4)
                       
                       insert_WHIP(cursor, new_sheet, iplEndStatement, season_totals_row, 13, exea, False,False, total_innings)
                       
                       insert_OBP(cursor, new_sheet, iplEndStatement, season_totals_row, 14, exea)
                       
                       insert_OPS(cursor, new_sheet, iplEndStatement, season_totals_row, 15, exea)
                       
                       insert_lead_off_out_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 16, exea, 65)
                       
                       insert_overall_strike_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 17, exea, 60)
                       
                       insert_baa(cursor,new_sheet,iplEndStatement,season_totals_row,18,exea)
                       
                       insert_baa_with_2_strikes(cursor, new_sheet, iplEndStatement, season_totals_row, 19, exea, 0.15)
                       
                       insert_baa_bip(cursor, new_sheet, iplEndStatement, season_totals_row, 20, exea)
                       
                       insert_freebases_count(cursor, new_sheet, iplEndStatement, season_totals_row, 21, exea)
                       
                       insert_strikeout_count(cursor, new_sheet, iplEndStatement, season_totals_row, 22, exea)
                       
                       insert_advantage_counts_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 23, exea)
                       
                       insert_disadvantage_counts_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 24, exea)
                                          
                       insert_strikeout_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 25, exea)
                       
                       insert_ground_ball_out_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 26, exea)
                       
                       insert_fly_ball_out_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 27, exea)
                       
                       insert_at_bat_win_rate(cursor, new_sheet, iplEndStatement, season_totals_row, 28, exea)
                       
                       insert_pitch_spread_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 29, exea)
                       
                       insert_pitch_spread_strike_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 30, exea)
                       
                       insert_pitch_spread_whiff_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 31, exea)  
                       
                       insert_pitch_spread_hits_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 32, exea)
                                          
                       bold_first_column_if_threshold(new_sheet, 5)
                       
                       insert_whip_by_inning_of_work(cursor, new_sheet, firstname, lastname, season_totals_row+4)
                       
                       insert_avg_peak_FB_velo_over_time_chart(cursor, new_sheet, firstname, lastname, season_totals_row+6)
                       
                       insert_movement_profile_chart(cursor, new_sheet, firstname, lastname, season_totals_row+6)
                       
                       insert_avg_pitch_velo_over_time(cursor, new_sheet, firstname, lastname, season_totals_row+6)
                       
                       adjust_formating(new_sheet, season_totals_row)
                  
           savebook(workbook, file_path, "Pitcher Logs Updated")
               
               
           
           
#wipes and updates pitchers logs
def wipe_and_up_pitchers_log(cursor,update_date,file_name):
           
           workbook=create_workbook(file_name)[0]
           file_path=create_workbook(file_name)[1]
         
           wipe (workbook)
           
           ##### Name
           cursor.execute("SELECT DISTINCT fname, lname FROM pitch_log_t WHERE pitch_id <>'0' Order By Fname,Lname ASC")
           data=cursor.fetchall()
           
           for i, (fname, lname) in enumerate(data,3):
               name = f"{fname} {lname}"
               sheetname=name 
               
                                    
               new_sheet=setup(sheetname,workbook,"Name",name,"Updated Date",update_date,"","",header_pos,pitcher_headers)
               

               query="SELECT DISTINCT date AS datea, date_n FROM pitch_log_t WHERE fname=%s AND lname = %s and pitch_id <> '0' ORDER BY date_n ASC"
               cursor.execute(query,(fname,lname))
               data=cursor.fetchall()
               
               total_pitch_count=0
               total_innings=0
               total_peak_velo=0
               pitchers=0
               
               
               for j, (datea) in enumerate(data,3):
                   dates=len(data)
                   datea =datea[0] if isinstance(datea, tuple) and datea else datea
                   new_sheet.cell(row=j, column=1, value=datea)
                   
                   exea=(datea,fname,lname)
                   iplEndStatement="WHERE date = %s AND fname = %s AND lname = %s AND pitch_id <> '0' "
                           
                   insert_oppo(cursor, new_sheet,iplEndStatement,j,2,exea)
                   
                   total_pitch_count+=insert_pitches_thrown(cursor, new_sheet, iplEndStatement, j, 3, exea)
                   
                   total_innings+=insert_pitches_per_inning(cursor, new_sheet, iplEndStatement, j, 4, exea,False)
                   
                   a, b=insert_peak_velo(cursor, new_sheet, iplEndStatement, j, 5, exea)
                  
                   total_peak_velo+=a
                   pitchers+=b
                               
                   insert_1st_pitch_strike_percentage(cursor, new_sheet, iplEndStatement, j, 6, exea, 60)
                   
                   insert_off_speed_strike_percentage(cursor, new_sheet, iplEndStatement, j, 7, exea, 50)
                   
                   insert_swing_and_miss_percentage(cursor, new_sheet,iplEndStatement, j, 8, exea, 25)
                        
                   insert_velo_range(cursor, new_sheet, iplEndStatement, j, 9, exea,False,False,False)        
                   
                   insert_chases(cursor, new_sheet, iplEndStatement, j, 10, exea, False, 0)
                   
                   insert_ahead_after_3_pitches_percentage(cursor,new_sheet,iplEndStatement,j,11,exea,60)
                   
                   insert_opponent_slugging_percentage(cursor, new_sheet, iplEndStatement, j, 12, exea, .4)
                   
                   insert_WHIP(cursor, new_sheet, iplEndStatement, j, 13, exea, False,False, 0)
                   
                   insert_OBP(cursor, new_sheet, iplEndStatement, j, 14, exea)
                   
                   insert_OPS(cursor, new_sheet, iplEndStatement, j, 15, exea)
                   
                   insert_lead_off_out_percentage(cursor, new_sheet, iplEndStatement, j, 16, exea, 65)
                   
                   insert_overall_strike_percentage(cursor, new_sheet, iplEndStatement, j, 17, exea, 60)
                   
                   insert_baa(cursor,new_sheet,iplEndStatement,j,18,exea)
                   
                   insert_baa_with_2_strikes(cursor, new_sheet, iplEndStatement, j, 19, exea, 0.15)
                   
                   insert_baa_bip(cursor, new_sheet, iplEndStatement, j, 20, exea)
                   
                   insert_freebases_count(cursor, new_sheet, iplEndStatement, j, 21, exea)
                   
                   insert_strikeout_count(cursor, new_sheet, iplEndStatement, j, 22, exea)
                           
                   insert_advantage_counts_percentage(cursor, new_sheet, iplEndStatement, j, 23, exea)
                   
                   insert_disadvantage_counts_percentage(cursor, new_sheet, iplEndStatement, j, 24, exea)
                                      
                   insert_strikeout_percentage(cursor, new_sheet, iplEndStatement, j, 25, exea)
                   
                   insert_ground_ball_out_percentage(cursor, new_sheet, iplEndStatement, j, 26, exea)
                   
                   insert_fly_ball_out_percentage(cursor, new_sheet, iplEndStatement, j, 27, exea)
                   
                   insert_at_bat_win_rate(cursor, new_sheet, iplEndStatement, j, 28, exea)
                   
                   insert_pitch_spread_percentage(cursor, new_sheet, iplEndStatement, j, 29, exea)
                   
                   insert_pitch_spread_strike_percentage(cursor, new_sheet, iplEndStatement, j, 30, exea)
                   
                   insert_pitch_spread_whiff_percentage(cursor, new_sheet, iplEndStatement, j, 31, exea)
                   
                   insert_pitch_spread_hits_percentage(cursor, new_sheet, iplEndStatement, j, 32, exea)
                   
               ######################## Player's Season Totals ####################################
               
               iplEndStatement="WHERE fname = %s AND lname = %s AND pitch_id <> '0' AND opponent <> 'Scrimmage' "
               exea=(fname,lname)
               
               # Calculate and insert team totals
               season_totals_row = dates + 4  # Assuming a gap of one row between individual pitchers and team totals
               
               new_sheet.cell(row=season_totals_row, column=1, value="Season Totals")
            
               ##### Total Pitch Count
               new_sheet.cell(row=season_totals_row, column=2, value="-")
        
               ##### Total Pitch Count
               new_sheet.cell(row=season_totals_row, column=3, value=total_pitch_count)
        
               ##### Average Pitches Per Inning
               if total_innings !=0:
                   new_sheet.cell(row=season_totals_row, column=4, value=round(total_pitch_count/total_innings,2))
               else:
                   new_sheet.cell(row=season_totals_row,column = 4, value = 0)
            
               #####  Average Peak Velocity
               if pitchers!=0:
                   new_sheet.cell(row=season_totals_row, column=5, value=round(total_peak_velo/pitchers,2))
               else:
                   new_sheet.cell(row=season_totals_row,column = 5, value = 0)
               
               
               insert_1st_pitch_strike_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 6, exea, 60)
               
               insert_off_speed_strike_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 7, exea, 50)
             
               insert_swing_and_miss_percentage(cursor, new_sheet,iplEndStatement, season_totals_row, 8, exea, 25)
                    
               insert_velo_range(cursor, new_sheet, iplEndStatement, season_totals_row, 9, exea,False,False,False)        
                       
               insert_chases(cursor, new_sheet, iplEndStatement, season_totals_row, 10, exea, False, total_innings)
               
               insert_ahead_after_3_pitches_percentage(cursor,new_sheet,iplEndStatement,season_totals_row,11,exea,60)
               
               insert_opponent_slugging_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 12, exea, .4)
               
               insert_WHIP(cursor, new_sheet, iplEndStatement, season_totals_row, 13, exea, False,False, total_innings)
              
               insert_OBP(cursor, new_sheet, iplEndStatement, season_totals_row, 14, exea)
              
               insert_OPS(cursor, new_sheet, iplEndStatement, season_totals_row, 15, exea)
              
               insert_lead_off_out_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 16, exea, 65)
              
               insert_overall_strike_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 17, exea, 60)
              
               insert_baa(cursor,new_sheet,iplEndStatement,season_totals_row,18,exea)
              
               insert_baa_with_2_strikes(cursor, new_sheet, iplEndStatement, season_totals_row, 19, exea, 0.15)
              
               insert_baa_bip(cursor, new_sheet, iplEndStatement, season_totals_row, 20, exea)
              
               insert_freebases_count(cursor, new_sheet, iplEndStatement, season_totals_row, 21, exea)
              
               insert_strikeout_count(cursor, new_sheet, iplEndStatement, season_totals_row, 22, exea)
              
               insert_advantage_counts_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 23, exea)
              
               insert_disadvantage_counts_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 24, exea)
                                 
               insert_strikeout_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 25, exea)
                  
               insert_ground_ball_out_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 26, exea)
              
               insert_fly_ball_out_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 27, exea)
              
               insert_at_bat_win_rate(cursor, new_sheet, iplEndStatement, season_totals_row, 28, exea)
               
               insert_pitch_spread_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 29, exea)
              
               insert_pitch_spread_strike_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 30, exea)
              
               insert_pitch_spread_whiff_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 31, exea)  
              
               insert_pitch_spread_hits_percentage(cursor, new_sheet, iplEndStatement, season_totals_row, 32, exea)
                                  
               bold_first_column_if_threshold(new_sheet, 5)
               
               insert_whip_by_inning_of_work(cursor, new_sheet, fname, lname, season_totals_row+4)
               
               insert_avg_peak_FB_velo_over_time_chart(cursor, new_sheet, fname, lname, season_totals_row+6)
               
               insert_movement_profile_chart(cursor, new_sheet, fname, lname, season_totals_row+6)
               
               insert_avg_pitch_velo_over_time(cursor, new_sheet, fname, lname, season_totals_row+6)
               
               adjust_formating(new_sheet, season_totals_row)
          
           savebook(workbook, file_path, "Pitcher Logs Updated")
        

#updates season log
def up_season_log(cursor,date,file_name):
            
    
           workbook=create_workbook(file_name)[0]
           file_path=create_workbook(file_name)[1]
           
           last_sheetname=workbook.worksheets[-1].title
           
           if (last_sheetname==date):
               del workbook[last_sheetname]
              
            
           sheetname=date
           
           new_sheet=setup(sheetname, workbook, "Updated Date", date, "", "", "", "", header_pos, season_game_headers)
           
           exea=()
           iplEndStatement="WHERE pitch_id <> '0' AND opponent <> 'Scrimmage' GROUP BY fname, lname ORDER BY fname,lname"
 
           ##### Name
           cursor.execute("SELECT DISTINCT fname, lname FROM pitch_log_t WHERE pitch_id <> '0' AND opponent <> 'Scrimmage' GROUP BY fname,lname Order By fname,lname")
           data=cursor.fetchall()
           
           for i, (fname, lname) in enumerate(data,3):
               full_name = f"{fname} {lname}"
               new_sheet.cell(row=i, column=1, value=full_name)
             
           total_pitch_count=insert_pitches_thrown(cursor, new_sheet, iplEndStatement, 0, 2, exea)
           
           total_innings=insert_pitches_per_inning(cursor, new_sheet, iplEndStatement, 0, 3, exea, True)
          
           total_peak_velo, pitchers=insert_peak_velo(cursor, new_sheet, iplEndStatement, 0, 4, exea) 
                       
           insert_1st_pitch_strike_percentage(cursor, new_sheet, iplEndStatement, 0, 5, exea, 60)
           
           insert_off_speed_strike_percentage(cursor, new_sheet, iplEndStatement, 0, 6, exea, 50)
         
           insert_swing_and_miss_percentage(cursor, new_sheet,iplEndStatement, 0, 7, exea, 25)
                
           insert_velo_range(cursor, new_sheet, iplEndStatement, 0, 8, exea,True,False,False)   
                   
           insert_chases(cursor, new_sheet, iplEndStatement, 0, 9, exea, True, 0)
           
           insert_ahead_after_3_pitches_percentage(cursor,new_sheet,iplEndStatement,0,10,exea,60)
           
           insert_opponent_slugging_percentage(cursor, new_sheet, iplEndStatement, 0, 11, exea, .4)
           
           insert_WHIP(cursor, new_sheet, iplEndStatement, 0, 12, exea, True,False, 0)
           
           insert_OBP(cursor, new_sheet, iplEndStatement, 0, 13, exea)
           
           insert_OPS(cursor, new_sheet, iplEndStatement, 0, 14, exea)
           
           insert_lead_off_out_percentage(cursor, new_sheet, iplEndStatement, 0, 15, exea, 65)
           
           insert_overall_strike_percentage(cursor, new_sheet, iplEndStatement, 0, 16, exea, 60)
           
           insert_baa(cursor,new_sheet,iplEndStatement,0,17,exea)
          
           insert_baa_with_2_strikes(cursor, new_sheet, iplEndStatement, 0, 18, exea, 0.15)
          
           insert_baa_bip(cursor, new_sheet, iplEndStatement, 0, 19, exea)
           
           insert_freebases_count(cursor, new_sheet, iplEndStatement, 0, 20, exea)
           
           insert_strikeout_count(cursor, new_sheet, iplEndStatement, 0, 21, exea)
           
           insert_advantage_counts_percentage(cursor, new_sheet, iplEndStatement, 0, 22, exea)
           
           insert_disadvantage_counts_percentage(cursor, new_sheet, iplEndStatement, 0, 23, exea)
                              
           insert_strikeout_percentage(cursor, new_sheet, iplEndStatement, 0, 24, exea)
           
           insert_ground_ball_out_percentage(cursor, new_sheet, iplEndStatement, 0, 25, exea)
           
           insert_fly_ball_out_percentage(cursor, new_sheet, iplEndStatement, 0, 26, exea)
           
           insert_at_bat_win_rate(cursor, new_sheet, iplEndStatement, 0, 27, exea)
           
           insert_pitch_spread_percentage(cursor, new_sheet, iplEndStatement, 0, 28, exea)
           
           insert_pitch_spread_strike_percentage(cursor, new_sheet, iplEndStatement, 0, 29, exea)
           
           insert_pitch_spread_whiff_percentage(cursor, new_sheet, iplEndStatement, 0, 30, exea)
           
           insert_pitch_spread_hits_percentage(cursor, new_sheet, iplEndStatement, 0, 31, exea)
           
           
               
           ################################## Team Totals ############################################# 
             
           exea=()
           iplEndStatement="WHERE pitch_id <> '0' AND opponent <> 'Scrimmage' "
            
           # Calculate and insert team totals
           team_totals_row = len(data) + 4  # Assuming a gap of one row between individual pitchers and team totals

           new_sheet.cell(row=team_totals_row, column=1, value="Team Totals")
           
           
           ##### Total Pitch Count
           new_sheet.cell(row=team_totals_row, column=2, value=total_pitch_count)
           
           ##### Average Pitches Per Inning
           if total_innings!=0:
               new_sheet.cell(row=team_totals_row, column=3, value=round(total_pitch_count/total_innings,2))
           else:
               new_sheet.cell(row=team_totals_row, column=3, value=0)
           
           #####  Average Peak Velocity
           if pitchers!=0:
               new_sheet.cell(row=team_totals_row, column=4, value=round(total_peak_velo/pitchers,2))
           else:
               new_sheet.cell(row=team_totals_row, column=4, value=0)
               
               
          
           insert_1st_pitch_strike_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 5, exea, 60)
           
           insert_off_speed_strike_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 6, exea, 50)
         
           insert_swing_and_miss_percentage(cursor, new_sheet,iplEndStatement, team_totals_row, 7, exea, 25)
                
           insert_velo_range(cursor, new_sheet, iplEndStatement, team_totals_row, 8, exea,True,False,False)        
                   
           insert_chases(cursor, new_sheet, iplEndStatement, team_totals_row, 9, exea, False, total_innings)
           
           insert_ahead_after_3_pitches_percentage(cursor,new_sheet,iplEndStatement,team_totals_row,10,exea,60)
           
           insert_opponent_slugging_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 11, exea, .4)
           
           insert_WHIP(cursor, new_sheet, iplEndStatement, team_totals_row, 12, exea, True,True, total_innings)
           
           insert_OBP(cursor, new_sheet, iplEndStatement, team_totals_row, 13, exea)
           
           insert_OPS(cursor, new_sheet, iplEndStatement, team_totals_row, 14, exea)
           
           insert_lead_off_out_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 15, exea, 65)
           
           insert_overall_strike_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 16, exea, 60)
           
           insert_baa(cursor,new_sheet,iplEndStatement,team_totals_row,17,exea)
          
           insert_baa_with_2_strikes(cursor, new_sheet, iplEndStatement, team_totals_row, 18, exea, 0.15)
          
           insert_baa_bip(cursor, new_sheet, iplEndStatement, team_totals_row, 19, exea)
           
           insert_freebases_count(cursor, new_sheet, iplEndStatement, team_totals_row, 20, exea)
           
           insert_strikeout_count(cursor, new_sheet, iplEndStatement, team_totals_row, 21, exea)
           
           insert_advantage_counts_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 22, exea)
           
           insert_disadvantage_counts_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 23, exea)
                              
           insert_strikeout_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 24, exea)
           
           insert_ground_ball_out_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 25, exea)
           
           insert_fly_ball_out_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 26, exea)
           
           insert_at_bat_win_rate(cursor, new_sheet, iplEndStatement, team_totals_row, 27, exea)
           
           insert_pitch_spread_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 28, exea)
           
           insert_pitch_spread_strike_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 29, exea)
           
           insert_pitch_spread_whiff_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 30, exea)
           
           insert_pitch_spread_hits_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 31, exea)
           
           bold_first_column_if_threshold(new_sheet, 5)
           
           adjust_formating(new_sheet, team_totals_row)
           
           
           savebook(workbook, file_path, "Season Log Updated")

           
#updates game log
def up_game_log(cursor,updated_date,file_name):

            workbook=create_workbook(file_name)[0]
            file_path=create_workbook(file_name)[1]
            
            last_sheetname=workbook.worksheets[-1].title
            
            
            last_updated=last_sheetname.split(" ")[0]
            
            last_up_query = "SELECT MAX(pitch_id) FROM pitch_log_T WHERE date=%s"
            cursor.execute(last_up_query,(last_updated,))
            data=cursor.fetchone()
            
            last_up=data[0]

            
            query="SELECT DISTINCT date AS datea,opponent as oppo, date_n AS new_date FROM pitch_log_t WHERE pitch_id<>0 AND pitch_id >%s ORDER BY date_n ASC"
            cursor.execute(query,(last_up,))
            data=cursor.fetchall()
            
            for j, (datea,oppo,new_date) in enumerate(data,3):
                
                datea = str(datea) if datea is not None else ""
                oppo = str(oppo) if oppo is not None else ""
                exea=(datea,oppo)
                
                iplEndStatement="WHERE date = %s AND opponent = %s AND pitch_id <> '0'  GROUP BY fname,lname ORDER BY fname,lname "
                

                sheetname=datea +" "+oppo
                
                new_sheet=setup(sheetname,workbook,"Date",datea,"Opponent",oppo,"Updated Date",updated_date,header_pos,season_game_headers)
                
                
                insert_names(cursor,new_sheet,iplEndStatement,0,1,exea)
                
                total_pitch_count=insert_pitches_thrown(cursor, new_sheet, iplEndStatement, 0, 2, exea)
                
                total_innings=insert_pitches_per_inning(cursor, new_sheet, iplEndStatement, 0, 3, exea,False)
                
                total_peak_velo,pitchers = insert_peak_velo(cursor, new_sheet, iplEndStatement, 0, 4, exea) 
                            
                insert_1st_pitch_strike_percentage(cursor, new_sheet, iplEndStatement, 0, 5, exea, 60)
                
                insert_off_speed_strike_percentage(cursor, new_sheet, iplEndStatement, 0, 6, exea, 50)
              
                insert_swing_and_miss_percentage(cursor, new_sheet,iplEndStatement, 0, 7, exea, 25)
                     
                insert_velo_range(cursor, new_sheet, iplEndStatement, 0, 8, exea,False,False,True)   
                        
                insert_chases(cursor, new_sheet, iplEndStatement, 0, 9, exea, False, 0)
                
                insert_ahead_after_3_pitches_percentage(cursor,new_sheet,iplEndStatement,0,10,exea,60)
                
                insert_opponent_slugging_percentage(cursor, new_sheet, iplEndStatement, 0, 11, exea, .4)
                
                insert_WHIP(cursor, new_sheet, iplEndStatement, 0, 12, exea, False,False, 0)
                
                insert_OBP(cursor, new_sheet, iplEndStatement, 0, 13, exea)
                
                insert_OPS(cursor, new_sheet, iplEndStatement, 0, 14, exea)
                
                insert_lead_off_out_percentage(cursor, new_sheet, iplEndStatement, 0, 15, exea, 65)
                
                insert_overall_strike_percentage(cursor, new_sheet, iplEndStatement, 0, 16, exea, 60)
                
                insert_baa(cursor,new_sheet,iplEndStatement,0,17,exea)
               
                insert_baa_with_2_strikes(cursor, new_sheet, iplEndStatement, 0, 18, exea, 0.15)
               
                insert_baa_bip(cursor, new_sheet, iplEndStatement, 0, 19, exea)
                
                insert_freebases_count(cursor, new_sheet, iplEndStatement, 0, 20, exea)
                
                insert_strikeout_count(cursor, new_sheet, iplEndStatement, 0, 21, exea)
                
                insert_advantage_counts_percentage(cursor, new_sheet, iplEndStatement, 0, 22, exea)
                
                insert_disadvantage_counts_percentage(cursor, new_sheet, iplEndStatement, 0, 23, exea)
                                   
                insert_strikeout_percentage(cursor, new_sheet, iplEndStatement, 0, 24, exea)
                
                insert_ground_ball_out_percentage(cursor, new_sheet, iplEndStatement, 0, 25, exea)
                
                insert_fly_ball_out_percentage(cursor, new_sheet, iplEndStatement, 0, 26, exea)
                
                insert_at_bat_win_rate(cursor, new_sheet, iplEndStatement, 0, 27, exea)
                
                insert_pitch_spread_percentage(cursor, new_sheet, iplEndStatement, 0, 28, exea)
                
                insert_pitch_spread_strike_percentage(cursor, new_sheet, iplEndStatement, 0, 29, exea)
                
                insert_pitch_spread_whiff_percentage(cursor, new_sheet, iplEndStatement, 0, 30, exea)
                
                insert_pitch_spread_hits_percentage(cursor, new_sheet, iplEndStatement, 0, 31, exea)
                        
                    
                ################################## Team Totals ############################################# 
                
                exea=(datea,oppo)
                iplEndStatement="WHERE date = %s AND opponent = %s AND pitch_id <> '0' "
                  
                # Calculate and insert team totals
                team_totals_row = pitchers + 4  # Assuming a gap of one row between individual pitchers and team totals

                new_sheet.cell(row=team_totals_row, column=1, value="Team Totals")
                
                ##### Total Pitch Count
                new_sheet.cell(row=team_totals_row, column=2, value=total_pitch_count)
                
                ##### Average Pitches Per Inning
                if total_innings!=0:
                    new_sheet.cell(row=team_totals_row, column=3, value=round(total_pitch_count/total_innings,2))
                else:
                    new_sheet.cell(row=team_totals_row, column=3, value=0)
                
                #####  Average Peak Velocity
                if pitchers!=0:
                    new_sheet.cell(row=team_totals_row, column=4, value=round(total_peak_velo/pitchers,2))
                else:
                    new_sheet.cell(row=team_totals_row, column=4, value=0)
                    
                
                insert_1st_pitch_strike_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 5, exea, 60)
                
                insert_off_speed_strike_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 6, exea, 50)
              
                insert_swing_and_miss_percentage(cursor, new_sheet,iplEndStatement, team_totals_row, 7, exea, 25)
                     
                insert_velo_range(cursor, new_sheet, iplEndStatement, team_totals_row, 8, exea,False,False,False)   
                        
                insert_chases(cursor, new_sheet, iplEndStatement, team_totals_row, 9, exea, False, total_innings)
                
                insert_ahead_after_3_pitches_percentage(cursor,new_sheet,iplEndStatement,team_totals_row,10,exea,60)
                
                insert_opponent_slugging_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 11, exea, .4)
                
                insert_WHIP(cursor, new_sheet, iplEndStatement, team_totals_row, 12, exea, False,False, total_innings)
                
                insert_OBP(cursor, new_sheet, iplEndStatement, team_totals_row, 13, exea)
                
                insert_OPS(cursor, new_sheet, iplEndStatement, team_totals_row, 14, exea)
                
                insert_lead_off_out_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 15, exea, 65)
                
                insert_overall_strike_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 16, exea, 60)
                
                insert_baa(cursor,new_sheet,iplEndStatement,team_totals_row,17,exea)
               
                insert_baa_with_2_strikes(cursor, new_sheet, iplEndStatement, team_totals_row, 18, exea, 0.15)
               
                insert_baa_bip(cursor, new_sheet, iplEndStatement, team_totals_row, 19, exea)
                
                insert_freebases_count(cursor, new_sheet, iplEndStatement, team_totals_row, 20, exea)
                
                insert_strikeout_count(cursor, new_sheet, iplEndStatement, team_totals_row, 21, exea)
                
                insert_advantage_counts_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 22, exea)
                
                insert_disadvantage_counts_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 23, exea)
                                   
                insert_strikeout_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 24, exea)
                
                insert_ground_ball_out_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 25, exea)
                
                insert_fly_ball_out_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 26, exea)
                
                insert_at_bat_win_rate(cursor, new_sheet, iplEndStatement, team_totals_row, 27, exea)
                
                insert_pitch_spread_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 28, exea)
                
                insert_pitch_spread_strike_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 29, exea)
                
                insert_pitch_spread_whiff_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 30, exea)
                
                insert_pitch_spread_hits_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 31, exea)
                  
                bold_first_column_if_threshold(new_sheet, 5)
                
                ##### INSERT PITCHES THAT GOT HIT #########
                
                insert_what_got_hit(cursor, new_sheet, team_totals_row+4, exea)
                
                adjust_formating(new_sheet, team_totals_row)
                
                
            savebook(workbook, file_path, "Game Log Updated")
            
            
           
#wipes and updates game log
def wipe_and_up_game_log(cursor,updated_date,file_name):

            workbook=create_workbook(file_name)[0]
            file_path=create_workbook(file_name)[1]
          
            wipe (workbook)
            
            query="SELECT DISTINCT date AS datea,opponent as oppo, date_n AS new_date FROM pitch_log_t WHERE pitch_id<>0 ORDER BY date_n ASC"
            cursor.execute(query)
            data=cursor.fetchall()
            
            for j, (datea,oppo,new_date) in enumerate(data,3):
                
                datea = str(datea) if datea is not None else ""
                oppo = str(oppo) if oppo is not None else ""
                exea=(datea,oppo)
                
                iplEndStatement="WHERE date = %s AND opponent = %s AND pitch_id <> '0'  GROUP BY fname,lname ORDER BY fname,lname "
                

                sheetname=datea +" "+oppo
                
                new_sheet=setup(sheetname,workbook,"Date",datea,"Opponent",oppo,"Updated Date",updated_date,header_pos,season_game_headers)
                
                
                insert_names(cursor,new_sheet,iplEndStatement,0,1,exea)
                
                total_pitch_count=insert_pitches_thrown(cursor, new_sheet, iplEndStatement, 0, 2, exea)
                
                total_innings=insert_pitches_per_inning(cursor, new_sheet, iplEndStatement, 0, 3, exea,False)
                
                total_peak_velo,pitchers = insert_peak_velo(cursor, new_sheet, iplEndStatement, 0, 4, exea) 
                            
                insert_1st_pitch_strike_percentage(cursor, new_sheet, iplEndStatement, 0, 5, exea, 60)
                
                insert_off_speed_strike_percentage(cursor, new_sheet, iplEndStatement, 0, 6, exea, 50)
              
                insert_swing_and_miss_percentage(cursor, new_sheet,iplEndStatement, 0, 7, exea, 25)
                     
                insert_velo_range(cursor, new_sheet, iplEndStatement, 0, 8, exea,False,False,True)   
                        
                insert_chases(cursor, new_sheet, iplEndStatement, 0, 9, exea, False, 0)
                
                insert_ahead_after_3_pitches_percentage(cursor,new_sheet,iplEndStatement,0,10,exea,60)
                
                insert_opponent_slugging_percentage(cursor, new_sheet, iplEndStatement, 0, 11, exea, .4)
                
                insert_WHIP(cursor, new_sheet, iplEndStatement, 0, 12, exea, False,False, 0)
                
                insert_OBP(cursor, new_sheet, iplEndStatement, 0, 13, exea)
                
                insert_OPS(cursor, new_sheet, iplEndStatement, 0, 14, exea)
                
                insert_lead_off_out_percentage(cursor, new_sheet, iplEndStatement, 0, 15, exea, 65)
                
                insert_overall_strike_percentage(cursor, new_sheet, iplEndStatement, 0, 16, exea, 60)
                
                insert_baa(cursor,new_sheet,iplEndStatement,0,17,exea)
               
                insert_baa_with_2_strikes(cursor, new_sheet, iplEndStatement, 0, 18, exea, 0.15)
               
                insert_baa_bip(cursor, new_sheet, iplEndStatement, 0, 19, exea)
                
                insert_freebases_count(cursor, new_sheet, iplEndStatement, 0, 20, exea)
                
                insert_strikeout_count(cursor, new_sheet, iplEndStatement, 0, 21, exea)
                
                insert_advantage_counts_percentage(cursor, new_sheet, iplEndStatement, 0, 22, exea)
                
                insert_disadvantage_counts_percentage(cursor, new_sheet, iplEndStatement, 0, 23, exea)
                                   
                insert_strikeout_percentage(cursor, new_sheet, iplEndStatement, 0, 24, exea)
                
                insert_ground_ball_out_percentage(cursor, new_sheet, iplEndStatement, 0, 25, exea)
                
                insert_fly_ball_out_percentage(cursor, new_sheet, iplEndStatement, 0, 26, exea)
                
                insert_at_bat_win_rate(cursor, new_sheet, iplEndStatement, 0, 27, exea)
                
                insert_pitch_spread_percentage(cursor, new_sheet, iplEndStatement, 0, 28, exea)
                
                insert_pitch_spread_strike_percentage(cursor, new_sheet, iplEndStatement, 0, 29, exea)
                
                insert_pitch_spread_whiff_percentage(cursor, new_sheet, iplEndStatement, 0, 30, exea)
                
                insert_pitch_spread_hits_percentage(cursor, new_sheet, iplEndStatement, 0, 31, exea)
                        
                    
                ################################## Team Totals ############################################# 
                
                exea=(datea,oppo)
                iplEndStatement="WHERE date = %s AND opponent = %s AND pitch_id <> '0' "
                  
                # Calculate and insert team totals
                team_totals_row = pitchers + 4  # Assuming a gap of one row between individual pitchers and team totals

                new_sheet.cell(row=team_totals_row, column=1, value="Team Totals")
                
                ##### Total Pitch Count
                new_sheet.cell(row=team_totals_row, column=2, value=total_pitch_count)
                
                ##### Average Pitches Per Inning
                if total_innings!=0:
                    new_sheet.cell(row=team_totals_row, column=3, value=round(total_pitch_count/total_innings,2))
                else:
                    new_sheet.cell(row=team_totals_row, column=3, value=0)
                
                #####  Average Peak Velocity
                if pitchers!=0:
                    new_sheet.cell(row=team_totals_row, column=4, value=round(total_peak_velo/pitchers,2))
                else:
                    new_sheet.cell(row=team_totals_row, column=4, value=0)
                    
                
                insert_1st_pitch_strike_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 5, exea, 60)
                
                insert_off_speed_strike_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 6, exea, 50)
              
                insert_swing_and_miss_percentage(cursor, new_sheet,iplEndStatement, team_totals_row, 7, exea, 25)
                     
                insert_velo_range(cursor, new_sheet, iplEndStatement, team_totals_row, 8, exea,False,False,False)   
                        
                insert_chases(cursor, new_sheet, iplEndStatement, team_totals_row, 9, exea, False, total_innings)
                
                insert_ahead_after_3_pitches_percentage(cursor,new_sheet,iplEndStatement,team_totals_row,10,exea,60)
                
                insert_opponent_slugging_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 11, exea, .4)
                
                insert_WHIP(cursor, new_sheet, iplEndStatement, team_totals_row, 12, exea, False,False, total_innings)
                
                insert_OBP(cursor, new_sheet, iplEndStatement, team_totals_row, 13, exea)
                
                insert_OPS(cursor, new_sheet, iplEndStatement, team_totals_row, 14, exea)
                
                insert_lead_off_out_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 15, exea, 65)
                
                insert_overall_strike_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 16, exea, 60)
                
                insert_baa(cursor,new_sheet,iplEndStatement,team_totals_row,17,exea)
               
                insert_baa_with_2_strikes(cursor, new_sheet, iplEndStatement, team_totals_row, 18, exea, 0.15)
               
                insert_baa_bip(cursor, new_sheet, iplEndStatement, team_totals_row, 19, exea)
                
                insert_freebases_count(cursor, new_sheet, iplEndStatement, team_totals_row, 20, exea)
                
                insert_strikeout_count(cursor, new_sheet, iplEndStatement, team_totals_row, 21, exea)
                
                insert_advantage_counts_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 22, exea)
                
                insert_disadvantage_counts_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 23, exea)
                                   
                insert_strikeout_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 24, exea)
                
                insert_ground_ball_out_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 25, exea)
                
                insert_fly_ball_out_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 26, exea)
                
                insert_at_bat_win_rate(cursor, new_sheet, iplEndStatement, team_totals_row, 27, exea)
                
                insert_pitch_spread_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 28, exea)
                
                insert_pitch_spread_strike_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 29, exea)
                
                insert_pitch_spread_whiff_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 30, exea)
                
                insert_pitch_spread_hits_percentage(cursor, new_sheet, iplEndStatement, team_totals_row, 31, exea)
                  
                bold_first_column_if_threshold(new_sheet, 5)
                
                ##### INSERT PITCHES THAT GOT HIT #########
                
                insert_what_got_hit(cursor, new_sheet, team_totals_row+4, exea)
                
                adjust_formating(new_sheet, team_totals_row)
                
                
            savebook(workbook, file_path, "Game Log Updated")



def main():
    try:
        connection = psycopg2.connect(
            dbname="ps1",
            user="pythoncon",
            password="password",
            host="18.217.248.114",
            port="5432"
        )
        
        with connection.cursor() as cursora:
            print ("Welcome to Pitching Log Creator. ")
            today = input("Enter today's date (PUT IT IN MM-DD-YYYY form): ")
        
            update = input("Would you like to update pitchers log, season log, game log, or all logs? (X to stop): ")
        
            while (update!='X'):
            
                print("")
                
                #update pitchers log
                if (update=='pitchers'):
                    up_pitchers_log(cursora,today,"Pitcher_Logs_2024_A.xlsx")
                
                #update season log
                if (update =='season'):
                    up_season_log(cursora,today,"Season_Logs_2024_A.xlsx")
                    
                #update game log
                if (update =='game'):
                    up_game_log(cursora,today,"Game_Logs_2024_A.xlsx")
                
                #update all logs
                if (update =='all'):
                    up_pitchers_log(cursora,today,"Pitcher_Logs_2024_A.xlsx")
                    up_season_log(cursora,today,"Season_Logs_2024_A.xlsx")
                    up_game_log(cursora,today,"Game_Logs_2024_A.xlsx")
                    update='X'
                
                #wipe and update all logs
                if (update =="wipe"):
                    wipe_and_up_pitchers_log(cursora,today,"Pitcher_Logs_2024_A.xlsx")
                    up_season_log(cursora,today,"Season_Logs_2024_A.xlsx")
                    wipe_and_up_game_log(cursora,today,"Game_Logs_2024_A.xlsx")
                    update='X'
                    
                if update not in ("pitchers","season","game","all","wipe","X"):
                    print("Invalid Entry")
                    
                print("")
                
                if (update!='X'):
                    update = input("Would you like to update game log, season log, pitchers log, or all logs? (X to stop): ")
                    if (update=='X'):
                        print("")
                
        
    except psycopg2.Error as e:
        # Handle database-related exceptions here
        print(f"Database error: {e}")

    except Exception as e:
        # Handle other exceptions here
        print(f"An unexpected error occurred: {e}")

    finally:
        # This block will be executed whether an exception occurs or not
        if connection:
            connection.close()
            print("Connection closed. GO LIONS!!!!")

if __name__ == "__main__":
    main()           